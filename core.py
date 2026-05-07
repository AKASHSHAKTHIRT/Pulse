from __future__ import annotations

import uuid
import sqlite3
import re
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta, time
from typing import Any, Dict, List, Optional, Tuple


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_PASSWORD = "Mach@123"

# =============================
# Helpers (time)
# =============================
def now() -> datetime:
    # Uses system date/time (PC clock)
    return datetime.now()


def dt_to_iso(dt: Optional[datetime]) -> Optional[str]:
    return dt.isoformat(timespec="seconds") if dt else None


def iso_to_dt(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s)
    except Exception:
        # fallback older format
        try:
            return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None


def fmt_dt(dt: Optional[datetime]) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""


def fmt_date(d: Optional[date]) -> str:
    return d.strftime("%Y-%m-%d") if d else ""


def parse_date(s: str) -> Optional[date]:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def parse_user_datetime(s: str) -> Optional[datetime]:
    """
    Parse user-entered datetime in common formats.
    Accepts:
      - YYYY-MM-DD HH:MM
      - YYYY-MM-DD HH:MM:SS
      - YYYY/MM/DD HH:MM
      - YYYY/MM/DD HH:MM:SS
    """
    s = (s or "").strip()
    if not s:
        return None
    fmts = [
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S",
    ]
    for f in fmts:
        try:
            return datetime.strptime(s, f)
        except Exception:
            pass
    return None



def short_id(n: int = 8) -> str:
    """Generate a short unique id for sessions."""
    return uuid.uuid4().hex[:n]


def seconds_to_hhmm(sec: int) -> str:
    sec = max(0, int(sec))
    h = sec // 3600
    m = (sec % 3600) // 60
    return f"{h:02d}:{m:02d}"


def seconds_to_hhmmss(sec: int) -> str:
    sec = max(0, int(sec))
    h = sec // 3600
    m = (sec % 3600) // 60
    s = sec % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def start_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())  # Monday


def daterange(d1: date, d2: date) -> List[date]:
    if d2 < d1:
        d1, d2 = d2, d1
    days = []
    cur = d1
    while cur <= d2:
        days.append(cur)
        cur += timedelta(days=1)
    return days


def clamp_dt(x: datetime, lo: datetime, hi: datetime) -> datetime:
    return max(lo, min(hi, x))


def split_session_by_day(start: datetime, end: datetime) -> Dict[date, int]:
    """Split [start, end] into per-day seconds."""
    out: Dict[date, int] = {}
    cur = start
    while cur.date() < end.date():
        midnight = datetime.combine(cur.date() + timedelta(days=1), time(0, 0, 0))
        out[cur.date()] = out.get(cur.date(), 0) + int((midnight - cur).total_seconds())
        cur = midnight
    out[end.date()] = out.get(end.date(), 0) + int((end - cur).total_seconds())
    return out
from openpyxl.styles import Alignment

def _merge_vertical(ws, col: int, r1: int, r2: int):
    """Merge ws cells vertically in one column if range has more than 1 row."""
    if r2 <= r1:
        return
    ws.merge_cells(start_row=r1, start_column=col, end_row=r2, end_column=col)
    c = ws.cell(row=r1, column=col)
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

def merge_employee_and_project(ws, start_row: int, end_row: int,
                               emp_col: int = 1, proj_col: int = 2):
    """
    Hierarchical merge:
      - Merge Employee column for consecutive same employees
      - Within each employee block, merge Project column for consecutive same projects
    """
    r = start_row
    while r <= end_row:
        emp = ws.cell(row=r, column=emp_col).value
        emp_start = r

        # find end of this employee block
        while r + 1 <= end_row and ws.cell(row=r + 1, column=emp_col).value == emp:
            r += 1
        emp_end = r

        # merge employee block
        if emp not in (None, ""):
            _merge_vertical(ws, emp_col, emp_start, emp_end)

        # within employee block, merge projects
        rr = emp_start
        while rr <= emp_end:
            proj = ws.cell(row=rr, column=proj_col).value
            proj_start = rr

            while rr + 1 <= emp_end and ws.cell(row=rr + 1, column=proj_col).value == proj:
                rr += 1
            proj_end = rr

            if proj not in (None, ""):
                _merge_vertical(ws, proj_col, proj_start, proj_end)

            rr += 1

        r += 1


# =============================
# Domain model
# =============================
TASK_CATEGORIES = ["Development", "HSIT", "LLRT", "Tools", "Support","SQA","Planning","Training"]

MODULE_CATEGORIES = ["DAU", "HDU", "CMU"]

# Backward-compatibility aliases (older builds used "DU" where the team now uses "HDU")
MODULE_ALIASES = {"DU": "HDU"}


def normalize_module(value: str) -> str:
    """Normalize module labels so summaries match (e.g., DU -> HDU)."""
    v = (value or "").strip()
    if not v:
        return ""
    v = v.upper()
    return MODULE_ALIASES.get(v, v)


@dataclass
class User:
    id: str
    name: str
    role: str  # "employee" | "admin"


@dataclass
class Employee:
    id: str
    name: str
    emp_code: str = ""   # login id like MGT-022
    laptop_brand: str = ""
    laptop_no: str = ""
    email: str = ""


@dataclass
class Project:
    id: str
    code: str
    name: str
    use_module: int = 1
    allowed_tasks: str = ""
    planned_hours: float = 0.0


@dataclass
class WorkSession:
    id: str
    employee_id: str
    project_id: str
    module: str
    task_category: str
    remark: str
    punch_in: datetime
    punch_out: Optional[datetime] = None

    @property
    def is_open(self) -> bool:
        return self.punch_out is None

    def duration_seconds(self) -> int:
        end = self.punch_out or now()
        return max(0, int((end - self.punch_in).total_seconds()))


@dataclass
class AppData:
    users: List[User] = field(default_factory=list)
    employees: List[Employee] = field(default_factory=list)
    projects: List[Project] = field(default_factory=list)
    sessions: List[WorkSession] = field(default_factory=list)

    def user_by_label(self, label: str) -> Optional[User]:
        name = label.split("(")[0].strip()
        role = label.split("(")[1].split(")")[0].strip() if "(" in label and ")" in label else ""
        return next((u for u in self.users if u.name == name and u.role == role), None)

    def employee_by_name(self, name: str) -> Optional[Employee]:
        return next((e for e in self.employees if e.name == name), None)

    def project_by_label(self, label: str) -> Optional[Project]:
        if "—" not in label:
            return None
        code = label.split("—", 1)[0].strip()
        return next((p for p in self.projects if p.code == code), None)

    def project_by_id(self, pid: str) -> Optional[Project]:
        return next((p for p in self.projects if p.id == pid), None)

    def get_open_session_for_employee(self, employee_id: str) -> Optional[WorkSession]:
        for s in reversed(self.sessions):
            if s.employee_id == employee_id and s.is_open:
                return s
        return None

    def total_closed_seconds_for_employee(self, employee_id: str, d_from: Optional[date] = None, d_to: Optional[date] = None) -> int:
        total = 0
        for s in self.sessions:
            if s.employee_id != employee_id or s.is_open:
                continue
            sd = s.punch_in.date()
            if d_from and sd < d_from:
                continue
            if d_to and sd > d_to:
                continue
            total += s.duration_seconds()
        return total

    def employee_by_code(self, code: str):
        code = (code or "").strip().upper()
        if not code:
            return None
        for e in self.employees:
            if (getattr(e, "emp_code", "") or "").strip().upper() == code:
                return e
        return None



# =============================
# Persistent storage (SQLite)
# =============================
class DataStore:
    """
    Local SQLite storage.
    - Keeps all sessions even if PC is restarted/off.
    - Open sessions remain open; timer continues using system time when app reopens.
    """

    def __init__(self, db_path: str = "time_tracker.db"):
        self.db_path = db_path
        self.conn = sqlite3.connect(self.db_path)
        self.conn.execute("PRAGMA foreign_keys = ON;")
        self._create_tables()
        self._ensure_employee_columns()
        self._ensure_session_columns()
        self._ensure_default_task_categories()
        # Keep legacy in-memory constant aligned with the DB-backed task list.
        try:
            TASK_CATEGORIES[:] = self.get_task_categories()
        except Exception:
            pass

    def _create_tables(self) -> None:
        cur = self.conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS users(
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                role TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS employees(
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                emp_code TEXT,
                laptop_brand TEXT,
                laptop_no TEXT,
                email TEXT DEFAULT ""
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS projects(
                id TEXT PRIMARY KEY,
                code TEXT NOT NULL,
                name TEXT NOT NULL,
                use_module INTEGER NOT NULL DEFAULT 1,
                allowed_tasks TEXT NOT NULL DEFAULT "",
                planned_hours REAL NOT NULL DEFAULT 0
            )
        """)

        # Per-project module planned hours
        cur.execute("""
            CREATE TABLE IF NOT EXISTS project_modules(
                project_id TEXT NOT NULL,
                module_name TEXT NOT NULL,
                planned_hours REAL NOT NULL DEFAULT 0,
                PRIMARY KEY(project_id, module_name),
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
        """)
                # Login credentials (per login id)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS project_soi_plan(
                project_id TEXT NOT NULL,
                task_name TEXT NOT NULL,
                soi_level INTEGER NOT NULL,
                planned_pct REAL NOT NULL DEFAULT 0,
                PRIMARY KEY(project_id, task_name),
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
        """)
                # Login credentials (per login id)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS credentials(
                login_id TEXT PRIMARY KEY,
                password TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS task_categories(
                name TEXT PRIMARY KEY
            )
        """)
        self.conn.commit()


        # --- Lightweight schema migration (safe for existing DBs) ---
        # Add new columns for per-project rules if missing.
        cur.execute("PRAGMA table_info(projects)")
        cols = {row[1] for row in cur.fetchall()}
        if "use_module" not in cols:
            cur.execute("ALTER TABLE projects ADD COLUMN use_module INTEGER NOT NULL DEFAULT 1")
        if "allowed_tasks" not in cols:
            cur.execute('ALTER TABLE projects ADD COLUMN allowed_tasks TEXT NOT NULL DEFAULT ""')
        if "planned_hours" not in cols:
            cur.execute("ALTER TABLE projects ADD COLUMN planned_hours REAL NOT NULL DEFAULT 0")

        self.conn.commit()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS sessions(
                id TEXT PRIMARY KEY,
                employee_id TEXT NOT NULL,
                project_id TEXT NOT NULL,
                module TEXT NOT NULL,
                task_category TEXT NOT NULL,
                remark TEXT NOT NULL,
                punch_in TEXT NOT NULL,
                punch_out TEXT,
                FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
        """)
        
        # Approvals: per employee per date-range
        cur.execute("""
            CREATE TABLE IF NOT EXISTS approvals(
                employee_id TEXT NOT NULL,
                date_from TEXT NOT NULL,
                date_to TEXT NOT NULL,
                approved INTEGER NOT NULL DEFAULT 0,
                approved_at TEXT,
                approved_by TEXT,
                PRIMARY KEY(employee_id, date_from, date_to),
                FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
            )
        """)
        self.conn.commit()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS alerts(
                id TEXT PRIMARY KEY,
                employee_id TEXT NOT NULL,
                message TEXT NOT NULL,
                created_at TEXT NOT NULL,
                seen INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(employee_id) REFERENCES employees(id) ON DELETE CASCADE
            )
        """)
        self.conn.commit()

    def ensure_credential(self, login_id: str, default_password: str = DEFAULT_PASSWORD) -> None:
        """Create credential row if missing."""
        login_id = (login_id or "").strip().upper()
        if not login_id:
            return
        cur = self.conn.cursor()
        cur.execute("SELECT 1 FROM credentials WHERE login_id=?", (login_id,))
        if cur.fetchone():
            return
        cur.execute("INSERT INTO credentials(login_id,password) VALUES(?,?)", (login_id, default_password))
        self.conn.commit()

    def _ensure_default_task_categories(self) -> None:
        """Seed built-in task categories ONLY if the table is completely empty.
        Never re-insert on every startup — that would undo user deletions."""
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) FROM task_categories")
        count = cur.fetchone()[0]
        if count == 0:
            for name in TASK_CATEGORIES:
                clean = (name or "").strip()
                if not clean:
                    continue
                cur.execute("INSERT OR IGNORE INTO task_categories(name) VALUES(?)", (clean,))
            self.conn.commit()

    def get_task_categories(self) -> List[str]:
        cur = self.conn.cursor()
        cur.execute("SELECT name FROM task_categories ORDER BY LOWER(name)")
        rows = [str(r[0]).strip() for r in (cur.fetchall() or []) if str(r[0]).strip()]
        return rows or list(TASK_CATEGORIES)

    def add_task_category(self, name: str) -> List[str]:
        clean = (name or "").strip()
        if not clean:
            raise ValueError("Name required")
        existing = self.get_task_categories()
        if clean.lower() in [t.lower() for t in existing]:
            raise ValueError(f"Task '{clean}' already exists")
        cur = self.conn.cursor()
        cur.execute("INSERT INTO task_categories(name) VALUES(?)", (clean,))
        self.conn.commit()
        updated = self.get_task_categories()
        TASK_CATEGORIES[:] = updated
        return updated

    def delete_task_category(self, name: str) -> List[str]:
        clean = (name or "").strip()
        if not clean:
            raise ValueError("Task name required")

        cur = self.conn.cursor()
        cur.execute("SELECT name FROM task_categories WHERE LOWER(name)=LOWER(?)", (clean,))
        row = cur.fetchone()
        if not row:
            raise ValueError(f"Task '{clean}' not found")

        actual_name = (row[0] or "").strip()

        # Delete the global task row.
        cur.execute("DELETE FROM task_categories WHERE LOWER(name)=LOWER(?)", (actual_name,))

        # Remove from per-project allowed_tasks CSV.
        cur.execute("SELECT id, allowed_tasks FROM projects")
        for pid, allowed in cur.fetchall() or []:
            tasks = [t.strip() for t in (allowed or "").split(',') if t.strip()]
            kept = [t for t in tasks if t.lower() != actual_name.lower()]
            if kept != tasks:
                cur.execute("UPDATE projects SET allowed_tasks=? WHERE id=?", (",".join(kept), pid))

        # Remove SOI mappings for this task across projects.
        cur.execute("DELETE FROM project_soi_plan WHERE LOWER(task_name)=LOWER(?)", (actual_name,))

        self.conn.commit()
        updated = self.get_task_categories()
        TASK_CATEGORIES[:] = updated
        return updated


    # =============================
    # Import: Hours from Excel (Employee Specific pivot)
    # =============================
    def _get_employee_id_by_name(self, employee_name: str) -> Optional[str]:
        employee_name = (employee_name or "").strip()
        if not employee_name:
            return None
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM employees WHERE LOWER(name)=LOWER(?) LIMIT 2",
            (employee_name,),
        )
        rows = cur.fetchall()
        if len(rows) != 1:
            # None (not found) OR ambiguous (duplicates)
            return None
        return rows[0][0]

    def _get_project_id_by_name(self, project_name: str) -> Optional[str]:
        project_name = (project_name or "").strip()
        if not project_name:
            return None
        cur = self.conn.cursor()
        cur.execute(
            "SELECT id FROM projects WHERE LOWER(name)=LOWER(?) LIMIT 2",
            (project_name,),
        )
        rows = cur.fetchall()
        if len(rows) != 1:
            return None
        return rows[0][0]

    @staticmethod
    def _hours_from_excel_cell(v) -> float:
        """Convert an Excel cell representing a duration into decimal hours."""
        if v is None:
            return 0.0

        # Excel time as datetime.time
        if isinstance(v, time):
            return (v.hour * 3600 + v.minute * 60 + v.second) / 3600.0

        # Excel timedelta
        if isinstance(v, timedelta):
            return v.total_seconds() / 3600.0

        # Numeric hours
        if isinstance(v, (int, float)):
            return float(v)

        # Strings like "2:30" or "2.5"
        if isinstance(v, str):
            s = v.strip()
            if not s:
                return 0.0
            if ":" in s:
                parts = s.split(":")
                try:
                    hh = int(parts[0])
                    mm = int(parts[1]) if len(parts) > 1 else 0
                    ss = int(parts[2]) if len(parts) > 2 else 0
                    return (hh * 3600 + mm * 60 + ss) / 3600.0
                except Exception:
                    return 0.0
            try:
                return float(s)
            except Exception:
                return 0.0

        return 0.0

    @staticmethod
    def _parse_employee_specific_header_to_date(header_value: str, default_year: int) -> Optional[date]:
        """Parses headers like 'Mon 05-Jan' into a date."""
        if not header_value:
            return None
        s = str(header_value).strip()
        m = re.search(r"(\d{1,2})-([A-Za-z]{3})", s)
        if not m:
            return None
        dd = int(m.group(1))
        mon_abbr = m.group(2).title()
        mon_map = {
            "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
            "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
        }
        mm = mon_map.get(mon_abbr)
        if not mm:
            return None
        try:
            return date(default_year, mm, dd)
        except Exception:
            return None

    def import_hours_from_employee_specific_excel(
        self,
        excel_path: str,
        sheet_name: str = "Employee Specific",
        default_module: str = "",
        default_task_category: str = "",
        start_time_hhmm: str = "09:00",
        replace_existing_for_day: bool = True,
    ) -> Dict[str, Any]:
        """Import pivot-style hours from the 'Employee Specific' sheet.

        The sheet format (as in Pulse_Report.xlsx):
        Sno | Employee | Project | Mon 05-Jan | Tue 06-Jan | ...
        Cell values are durations (Excel time / timedelta).

        We convert each non-zero duration into a synthetic closed session with:
          punch_in  = <date> <start_time_hhmm>
          punch_out = punch_in + duration

        Returns summary dict with inserted/skipped/errors.
        """
        summary: Dict[str, Any] = {
            "inserted": 0,
            "skipped": 0,
            "errors": [],
        }

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            summary["errors"].append(f"Failed to open Excel: {e}")
            return summary

        if sheet_name not in wb.sheetnames:
            summary["errors"].append(f"Sheet not found: {sheet_name}")
            return summary

        ws = wb[sheet_name]

        # Determine default year from file metadata if possible
        try:
            default_year = (getattr(wb.properties, "created", None) or datetime.now()).year
        except Exception:
            default_year = datetime.now().year

        # -----------------------------
        # Detect columns from header row
        # -----------------------------
        # Expected format (your latest sheet):
        # Sno | Employee | Project | Module | Mon 05-Jan | Tue 06-Jan | ... | Total Hours
        # Some cells may be merged (Employee), so we also forward-fill employee/module.

        headers: List[Tuple[int, str]] = []  # (col, header_str)
        c = 1
        while True:
            hv = ws.cell(1, c).value
            if hv is None:
                break
            headers.append((c, str(hv).strip()))
            c += 1

        # Find core columns by name
        def _find_col(*names: str) -> Optional[int]:
            wanted = {n.strip().lower() for n in names if n}
            for col, h in headers:
                if h.strip().lower() in wanted:
                    return col
            return None

        emp_col = _find_col("employee", "employee name", "name") or 2
        proj_col = _find_col("project", "project name") or 3
        module_col = _find_col("module")  # optional
        task_col = _find_col("task", "task category")  # optional

        # Parse date headers from row 1 (any column whose header looks like "Mon 05-Jan")
        date_cols: List[Tuple[int, date]] = []
        for col, h in headers:
            d = self._parse_employee_specific_header_to_date(h, default_year)
            if d:
                date_cols.append((col, d))

        # Keep date columns in ascending column order
        date_cols.sort(key=lambda x: x[0])

        if not date_cols:
            summary["errors"].append("No date columns detected in header row.")
            return summary

        # Start time
        try:
            st_h, st_m = [int(x) for x in start_time_hhmm.split(":", 1)]
            start_t = time(st_h, st_m, 0)
        except Exception:
            start_t = time(9, 0, 0)

        cur = self.conn.cursor()
        inserted = 0
        skipped = 0

        last_emp_name_s = ""
        # Data rows start at 2
        for r in range(2, ws.max_row + 1):
            emp_name = ws.cell(r, emp_col).value
            proj_name = ws.cell(r, proj_col).value
            mod_val = ws.cell(r, module_col).value if module_col else None
            task_val = ws.cell(r, task_col).value if task_col else None

            # Completely blank row
            if emp_name is None and proj_name is None and mod_val is None and task_val is None:
                continue

            emp_name_s = (str(emp_name).strip() if emp_name is not None else "")
            proj_name_s = (str(proj_name).strip() if proj_name is not None else "")
            module_s = (str(mod_val).strip() if mod_val is not None else "")
            task_s = (str(task_val).strip() if task_val is not None else "")

            # Forward-fill merged Employee / Module
            if not emp_name_s and last_emp_name_s:
                emp_name_s = last_emp_name_s
            else:
                last_emp_name_s = emp_name_s or last_emp_name_s

            # IMPORTANT: Do NOT forward-fill Module/Task.
            # In the Excel, Module/Task may be intentionally blank for a row.
            # If blank, we keep it blank (or fall back to defaults during insert).

            # Skip totals rows like "EMPLOYEE TOTAL"
            if (proj_name_s or "").strip().upper() in {"EMPLOYEE TOTAL", "TOTAL"}:
                continue

            if not emp_name_s or not proj_name_s:
                continue

            employee_id = self._get_employee_id_by_name(emp_name_s)
            if not employee_id:
                skipped += 1
                summary["errors"].append(f"Row {r}: Employee not found/ambiguous: '{emp_name_s}'")
                continue

            project_id = self._get_project_id_by_name(proj_name_s)
            if not project_id:
                skipped += 1
                summary["errors"].append(f"Row {r}: Project not found/ambiguous: '{proj_name_s}'")
                continue

            # Validate/choose module for this row
            # Only fall back to default_module when the sheet has NO Module column at all.
            has_module_col = module_col is not None
            default_module_n = normalize_module(default_module)
            row_module = normalize_module(module_s)

            # If the sheet doesn't have a Module column, allow UI/default to drive it.
            if not has_module_col and not row_module:
                row_module = default_module_n

            # If the sheet has a Module column and it's blank, keep it blank.
            # If it's a custom module name, keep it (don't silently rewrite hours).

            # Choose task for this row
            # If the sheet contains a Task column and the cell is blank, we KEEP it blank.
            # Only fall back to default_task_category when the sheet has NO Task column at all.
            has_task_col = task_col is not None
            row_task = (task_s or "").strip()
            if not row_task:
                row_task = "" if has_task_col else (default_task_category or "Development")
            elif row_task not in TASK_CATEGORIES:
                row_task = default_task_category or "Development"

            for c, d in date_cols:
                hours = self._hours_from_excel_cell(ws.cell(r, c).value)
                if hours <= 0.0:
                    continue

                punch_in_dt = datetime.combine(d, start_t)
                punch_out_dt = punch_in_dt + timedelta(hours=hours)

                if replace_existing_for_day:
                    day_start = datetime.combine(d, time(0, 0, 0))
                    day_end = day_start + timedelta(days=1)
                    cur.execute(
                        """
                        DELETE FROM sessions
                        WHERE employee_id=? AND project_id=?
                          AND punch_in >= ? AND punch_in < ?
                        """,
                        (employee_id, project_id, day_start.isoformat(), day_end.isoformat()),
                    )

                # Use existing id helper (gen_id was referenced in older code).
                sid = short_id()
                cur.execute(
                    """
                    INSERT INTO sessions(id,employee_id,project_id,module,task_category,remark,punch_in,punch_out)
                    VALUES (?,?,?,?,?,?,?,?)
                    """,
                    (
                        sid,
                        employee_id,
                        project_id,
                        (row_module or ""),
                        row_task,
                        "",
                        punch_in_dt.isoformat(),
                        punch_out_dt.isoformat(),
                    ),
                )
                inserted += 1

        self.conn.commit()
        summary["inserted"] = inserted
        summary["skipped"] = skipped
        return summary

    def set_password(self, login_id: str, new_password: str) -> None:
        """Admin reset / user set password."""
        login_id = (login_id or "").strip().upper()
        new_password = (new_password or "").strip()
        if not login_id:
            raise ValueError("Missing login id")
        if len(new_password) < 4:
            raise ValueError("Password must be at least 4 characters.")
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO credentials(login_id,password)
            VALUES(?,?)
            ON CONFLICT(login_id) DO UPDATE SET password=excluded.password
        """, (login_id, new_password))
        self.conn.commit()

    def verify_password(self, login_id: str, password: str) -> bool:
        """Return True if login_id exists and password matches.
        If credential row is missing, fallback to DEFAULT_PASSWORD.
        """
        login_id = (login_id or "").strip().upper()
        password = password or ""
        if not login_id:
            return False
        cur = self.conn.cursor()
        cur.execute("SELECT password FROM credentials WHERE login_id=?", (login_id,))
        row = cur.fetchone()
        if not row:
            # fallback: allow old behavior
            return password == DEFAULT_PASSWORD
        return password == (row[0] or "")

    def delete_credential(self, login_id: str) -> None:
        login_id = (login_id or "").strip().upper()
        if not login_id:
            return
        cur = self.conn.cursor()
        cur.execute("DELETE FROM credentials WHERE login_id=?", (login_id,))
        self.conn.commit()

    # ------------------------------
    # Auto-close unended tasks (midnight rule)
    # ------------------------------
    def _resolve_employee_db_id(self, key: str) -> Optional[str]:
        """Resolve a login key (emp_code like MGT-001 OR employee table id like e02) to employees.id."""
        key = (key or "").strip()
        if not key:
            return None
        cur = self.conn.cursor()
        # direct id
        cur.execute("SELECT id FROM employees WHERE id = ?", (key,))
        row = cur.fetchone()
        if row:
            return row[0]
        # emp_code
        cur.execute("SELECT id FROM employees WHERE UPPER(COALESCE(emp_code,'')) = ?", (key.upper(),))
        row = cur.fetchone()
        return row[0] if row else None
    def auto_zero_unended_tasks(self, employee_key: str) -> List[str]:
        """If employee has any open session from a previous day (not ended by 12:00 AM),
        set it to 0 hours (punch_out = punch_in), add an alert, and return any unseen alerts.
        Alert message includes Project name and Task category.
        """
        eid = self._resolve_employee_db_id(employee_key)
        if not eid:
            return []

        today = now().date()
        cur = self.conn.cursor()

        # Find open sessions (no punch_out yet)
        cur.execute(
            """
            SELECT s.id, s.punch_in, s.remark, s.project_id, s.module, s.task_category,
                   COALESCE(p.name, '') AS project_name
            FROM sessions s
            LEFT JOIN projects p ON p.id = s.project_id
            WHERE s.employee_id=? AND s.punch_out IS NULL
            """,
            (eid,),
        )
        rows = cur.fetchall() or []

        for sid, pin_s, remark, project_id, module, task_category, project_name in rows:
            pin_dt = iso_to_dt(pin_s)
            if not pin_dt:
                continue

            # only auto-close sessions from previous day(s)
            if pin_dt.date() < today:
                # Auto-close to zero hours (punch_out = punch_in)
                new_remark = (remark or "").strip()
                tag = "[AUTO-CLOSED: not ended by 12:00 AM]"
                if tag not in new_remark:
                    new_remark = (new_remark + " " + tag).strip()

                cur.execute(
                    "UPDATE sessions SET punch_out=?, remark=? WHERE id=?",
                    (pin_s, new_remark, sid),
                )

                proj = (project_name or "").strip() or (project_id or "").strip() or "-"
                mod = (module or "").strip() or "-"
                cat = (task_category or "").strip() or "-"

                msg = (
                    f"On {pin_dt.strftime('%Y-%m-%d')}, your task was not ended by 12:00 AM and was auto-closed as 0 hours.\n"
                    f"Project: {proj}\n"
                    f"Module: {mod}\n"
                    f"Category: {cat}"
                )

                cur.execute(
                    "INSERT INTO alerts(id, employee_id, message, created_at, seen) VALUES (?,?,?,?,0)",
                    (short_id(), eid, msg, dt_to_iso(now())),
                )

        self.conn.commit()

        # Return unseen alerts and mark them as seen
        cur.execute(
            "SELECT id, message FROM alerts WHERE employee_id=? AND seen=0 ORDER BY created_at ASC",
            (eid,),
        )
        msgs = [r[1] for r in (cur.fetchall() or [])]

        if msgs:
            cur.execute(
                "UPDATE alerts SET seen=1 WHERE employee_id=? AND seen=0",
                (eid,),
            )
            self.conn.commit()

        return msgs




    def _ensure_employee_columns(self) -> None:
        """Add new employee columns if DB was created by an older version."""
        cur = self.conn.cursor()
        cur.execute("PRAGMA table_info(employees)")
        cols = {row[1] for row in cur.fetchall()}

        for col, typ in [("emp_code", "TEXT"), ("laptop_brand", "TEXT"), ("laptop_no", "TEXT"), ("email", "TEXT")]:
            if col not in cols:
                cur.execute(f"ALTER TABLE employees ADD COLUMN {col} {typ}")

        self.conn.commit()

    def _ensure_session_columns(self) -> None:
        """Add new session columns if DB was created by an older version."""
        cur = self.conn.cursor()
        cur.execute("PRAGMA table_info(sessions)")
        cols = {row[1] for row in cur.fetchall()}
        if "module" not in cols:
            cur.execute("ALTER TABLE sessions ADD COLUMN module TEXT NOT NULL DEFAULT 'N/A'")
        self.conn.commit()


    def seed_demo(self) -> None:
        """Seed initial demo/master data.

        This app is used in production-like workflows where admins can add/delete
        projects and employees. If we keep re-seeding demo rows after the database
        has been customized, we can hit FOREIGN KEY errors (e.g., demo sessions
        referencing deleted demo projects like p1/p2/p3).

        Rule:
        - Only seed demo projects + demo sessions when BOTH tables are empty.
        - Always ensure the admin user exists.
        - Employees are only seeded when the employees table is empty.
        """
        cur = self.conn.cursor()

        # Ensure admin user exists (always)
        cur.execute(
            "INSERT OR REPLACE INTO users(id,name,role) VALUES (?,?,?)",
            ("u_admin", "Krishna", "admin"),
        )
        self.conn.commit()

        # If the DB was already customized (any employees/projects/sessions exist),
        # do NOT seed demo rows again. This prevents FOREIGN KEY errors such as:
        # - employees exist but demo employee ids (e02/e03) don't
        # - projects exist but demo project ids (p1/p2/p3) don't
        cur.execute("SELECT COUNT(*) FROM employees")
        employees_count = (cur.fetchone() or [0])[0]
        cur.execute("SELECT COUNT(*) FROM projects")
        projects_count = (cur.fetchone() or [0])[0]
        cur.execute("SELECT COUNT(*) FROM sessions")
        sessions_count = (cur.fetchone() or [0])[0]
        if employees_count > 0 or projects_count > 0 or sessions_count > 0:
            return

        # Seed employees (DB is empty here by rule above)
        employees = [
            ('e02', 'Anitha', 'MGT-002'),
            ('e03', 'Deepan Raj', 'MGT-005'),
            ('e04', 'Mohan Raj S', 'MGT-003'),
            ('e05', 'Akash Shakthi', 'MGT-017'),
            ('e06', 'Dhanush Priyan', ''),
            ('e07', 'Hemagiri', 'MGT-016'),
            ('e08', 'Kaviya', 'MGT-013'),
            ('e09', 'Madhulatha', 'MGT-009'),
            ('e10', 'Manasa', 'MGT-008'),
            ('e11', 'Monisha', 'MGT-007'),
            ('e12', 'Swathi', 'MGT-015'),
            ('e13', 'Swedha', ''),
            ('e14', 'Syed Saiful', 'MGT-012'),
            ('e15', 'Triveni', 'MGT-011'),
            ('e16', 'Yugandhar', 'MGT-004'),
            ('e17', 'Abinithi', 'MGT-020'),
            ('e18', 'Bharath', 'MGT-022'),
            ('e19', 'Sakthivel', 'MGT-021'),
            ('e20', 'Yugandhar R', 'MGT-019'),
        ]
        cur.executemany(
            "INSERT OR REPLACE INTO employees(id,name,emp_code) VALUES (?,?,?)",
            employees
        )
        self.conn.commit()

        # Seed demo projects (DB is empty here by rule above)
        projects = [
            ("p1", "PROJ-001", "E-Commerce Website"),
            ("p2", "PROJ-002", "Internal HR Tool"),
            ("p3", "PROJ-003", "Client Support"),
        ]
        cur.executemany("INSERT INTO projects(id,code,name) VALUES (?,?,?)", projects)

        # Seed demo sessions (DB is empty here by rule above)
        base = now() - timedelta(days=1)

        def add(emp_id: str, proj_id: str, cat: str, start: datetime, minutes: int, remark: str = ""):
            sid = short_id()
            pin = dt_to_iso(start)
            pout = dt_to_iso(start + timedelta(minutes=minutes))
            cur.execute(
                "INSERT INTO sessions(id,employee_id,project_id,module,task_category,remark,punch_in,punch_out) VALUES (?,?,?,?,?,?,?,?)",
                (sid, emp_id, proj_id, "", cat, remark, pin, pout),
            )

        add("e02", "p1", "Development", base.replace(hour=10, minute=0, second=0, microsecond=0), 55)
        add("e02", "p1", "Development", base.replace(hour=11, minute=10, second=0, microsecond=0), 40)
        add("e02", "p3", "Support", base.replace(hour=15, minute=0, second=0, microsecond=0), 35, "Helped client on urgent bug")
        add("e02", "p1", "HSIT", base.replace(hour=12, minute=30, second=0, microsecond=0), 70)
        add("e02", "p2", "Tools", base.replace(hour=16, minute=0, second=0, microsecond=0), 45)

        # Open (running) session persisted
        sid = short_id()
        cur.execute(
            "INSERT INTO sessions(id,employee_id,project_id,module,task_category,remark,punch_in,punch_out) VALUES (?,?,?,?,?,?,?,NULL)",
            (sid, "e03", "p3", "DAU", "LLRT", "", dt_to_iso(now() - timedelta(minutes=12))),
        )
        self.conn.commit()

    def add_employee(self, name: str, emp_code: str, laptop_brand: str = "", laptop_no: str = "", email: str = "") -> str:
        """Add a new employee (admin) and return the new employee DB id."""
        name = (name or "").strip()
        emp_code = (emp_code or "").strip().upper()
        laptop_brand = (laptop_brand or "").strip()
        laptop_no = (laptop_no or "").strip()
        email     = (email or "").strip().lower()

        if not name:
            raise ValueError("Employee name is required.")
        if not emp_code:
            raise ValueError("Employee ID is required (e.g., MGT-022).")

        cur = self.conn.cursor()

        # Unique emp_code (case-insensitive)
        cur.execute("SELECT id FROM employees WHERE UPPER(COALESCE(emp_code,'')) = ?", (emp_code,))
        if cur.fetchone():
            raise ValueError(f"Employee ID already exists: {emp_code}")

        # Next id like e21
        cur.execute("SELECT id FROM employees")
        ids = [r[0] for r in cur.fetchall()]
        max_n = 0
        for eid in ids:
            if isinstance(eid, str) and eid.startswith("e") and eid[1:].isdigit():
                max_n = max(max_n, int(eid[1:]))
        new_id = f"e{max_n+1:02d}" if max_n > 0 else "e01"
        if new_id in ids:
            new_id = "e" + short_id()

        cur.execute(
            "INSERT INTO employees(id,name,emp_code,laptop_brand,laptop_no,email) VALUES (?,?,?,?,?,?)",
            (new_id, name, emp_code, laptop_brand, laptop_no, email),
        )
        self.conn.commit()
        return new_id

    def delete_employee(self, employee_id: str) -> None:
        """Delete an employee and ALL their sessions."""
        employee_id = (employee_id or "").strip()
        if not employee_id:
            raise ValueError("Missing employee id")

        cur = self.conn.cursor()
        cur.execute("DELETE FROM sessions WHERE employee_id = ?", (employee_id,))
        cur.execute("DELETE FROM employees WHERE id = ?", (employee_id,))
        self.conn.commit()


    def update_employee(self, employee_id: str, name: str, emp_code: str, email: str = "") -> None:
        """Update employee name / emp_code (admin). Also migrates credentials login_id if emp_code changes."""
        employee_id = (employee_id or "").strip()
        name = (name or "").strip()
        emp_code = (emp_code or "").strip().upper()

        if not employee_id:
            raise ValueError("Missing employee id")
        if not name:
            raise ValueError("Employee name is required.")
        if not emp_code:
            raise ValueError("Employee ID is required.")

        cur = self.conn.cursor()

        # Fetch old code
        cur.execute("SELECT COALESCE(emp_code,'') FROM employees WHERE id=?", (employee_id,))
        row = cur.fetchone()
        if not row:
            raise ValueError("Employee not found")
        old_code = (row[0] or "").strip().upper()

        # Ensure emp_code unique among employees (excluding this employee)
        cur.execute("SELECT id FROM employees WHERE UPPER(COALESCE(emp_code,''))=? AND id<>?", (emp_code, employee_id))
        if cur.fetchone():
            raise ValueError(f"Employee ID already exists: {emp_code}")

        email = (email or "").strip().lower()
        # Update employee
        cur.execute("UPDATE employees SET name=?, emp_code=?, email=? WHERE id=?", (name, emp_code, email, employee_id))

        # Migrate credentials login_id if exists
        if old_code and old_code != emp_code:
            cur.execute("SELECT login_id FROM credentials WHERE login_id=?", (old_code,))
            old_cred = cur.fetchone()
            if old_cred:
                # If new code already has credential, drop old to avoid conflict
                cur.execute("SELECT login_id FROM credentials WHERE login_id=?", (emp_code,))
                new_cred = cur.fetchone()
                if new_cred:
                    cur.execute("DELETE FROM credentials WHERE login_id=?", (old_code,))
                else:
                    cur.execute("UPDATE credentials SET login_id=? WHERE login_id=?", (emp_code, old_code))

        self.conn.commit()
        return

    def add_project(self, code: str, name: str, use_module: bool = True, allowed_tasks: str = "", planned_hours: float = 0.0) -> str:
        """Create a new project. Returns the new project's id."""
        code = (code or "").strip().upper()
        name = (name or "").strip()
        if not code or not name:
            raise ValueError("Project code and name are required")
        pid = str(uuid.uuid4())
        cur = self.conn.cursor()
        # prevent duplicate project codes
        cur.execute("SELECT 1 FROM projects WHERE UPPER(code)=UPPER(?)", (code,))
        if cur.fetchone():
            raise ValueError(f"Project code already exists: {code}")
        cur.execute("INSERT INTO projects(id, code, name, use_module, allowed_tasks, planned_hours) VALUES(?,?,?,?,?,?)", (pid, code, name, 1 if use_module else 0, (allowed_tasks or "").strip(), float(planned_hours or 0.0)))
        self.conn.commit()
        return pid

    def update_project_plans(self, project_id: str, planned_hours: float, use_module: Optional[bool] = None, allowed_tasks: Optional[str] = None) -> None:
        """Update editable project plan fields (planned hours, module requirement, allowed tasks)."""
        cur = self.conn.cursor()

        # Keep existing values if not provided
        if use_module is None or allowed_tasks is None:
            row = cur.execute("SELECT use_module, allowed_tasks FROM projects WHERE id=?", (project_id,)).fetchone()
            if not row:
                raise ValueError("Project not found.")
            if use_module is None:
                use_module = bool(row[0])
            if allowed_tasks is None:
                allowed_tasks = row[1] or ""

        cur.execute(
            "UPDATE projects SET planned_hours=?, use_module=?, allowed_tasks=? WHERE id=?",
            (float(planned_hours or 0.0), 1 if use_module else 0, allowed_tasks or "", project_id),
        )
        self.conn.commit()

    # -----------------------------
    # Project Module Plans (planned hours per module)
    # -----------------------------
    def set_project_modules(self, project_id: str, modules: List[Tuple[str, float]]) -> None:
        """Overwrite module planned hours for a project.

        modules: [(module_name, planned_hours), ...]
        """
        project_id = (project_id or "").strip()
        if not project_id:
            return
        cur = self.conn.cursor()
        cur.execute("DELETE FROM project_modules WHERE project_id=?", (project_id,))
        for name, hrs in (modules or []):
            name = (name or "").strip()
            if not name:
                continue
            try:
                hrs_f = float(hrs or 0.0)
            except Exception:
                hrs_f = 0.0
            cur.execute(
                "INSERT OR REPLACE INTO project_modules(project_id, module_name, planned_hours) VALUES (?,?,?)",
                (project_id, normalize_module(name) or name, hrs_f),
            )
        self.conn.commit()

    def get_project_modules(self, project_id: str) -> List[Tuple[str, float]]:
        project_id = (project_id or "").strip()
        if not project_id:
            return []
        cur = self.conn.cursor()
        cur.execute(
            "SELECT module_name, planned_hours FROM project_modules WHERE project_id=? ORDER BY module_name",
            (project_id,),
        )
        return [(normalize_module(r[0]) or str(r[0] or ""), float(r[1] or 0.0)) for r in (cur.fetchall() or [])]

    def get_project_modules_map(self) -> Dict[Tuple[str, str], float]:
        """Return {(project_name, module_name): planned_hours}."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT p.name, pm.module_name, pm.planned_hours
            FROM project_modules pm
            JOIN projects p ON p.id = pm.project_id
            """
        )
        out: Dict[Tuple[str, str], float] = {}
        for pname, mname, ph in (cur.fetchall() or []):
            out[(str(pname), normalize_module(str(mname)) or str(mname))] = float(ph or 0.0)
        return out

    # -----------------------------
    # SOI Plan (tasks grouped under SOI-1..SOI-4 with planned %)
    # -----------------------------
    def set_project_soi_plan(self, project_id: str, rows: List[Tuple[str, int, float]]) -> None:
        """Overwrite SOI plan for a project.

        rows: [(task_name, soi_level(1..4), planned_pct), ...]
        """
        project_id = (project_id or "").strip()
        if not project_id:
            return
        cur = self.conn.cursor()
        cur.execute("DELETE FROM project_soi_plan WHERE project_id=?", (project_id,))
        for task, soi, pct in (rows or []):
            task = (task or "").strip()
            if not task:
                continue
            try:
                soi_i = int(soi)
            except Exception:
                continue
            if soi_i not in (1, 2, 3, 4):
                continue
            try:
                pct_f = float(pct or 0.0)
            except Exception:
                pct_f = 0.0
            cur.execute(
                "INSERT OR REPLACE INTO project_soi_plan(project_id, task_name, soi_level, planned_pct) VALUES (?,?,?,?)",
                (project_id, task, soi_i, pct_f),
            )
        self.conn.commit()

    def get_project_soi_plan(self, project_id: str) -> List[Tuple[str, int, float]]:
        """Return [(task_name, soi_level, planned_pct), ...]"""
        project_id = (project_id or "").strip()
        if not project_id:
            return []
        cur = self.conn.cursor()
        cur.execute(
            "SELECT task_name, soi_level, planned_pct FROM project_soi_plan WHERE project_id=? ORDER BY soi_level, task_name",
            (project_id,),
        )
        return [(row[0], int(row[1] or 1), float(row[2] or 0.0)) for row in cur.fetchall()]

    def get_soi_plan_by_project_name(self) -> Dict[Tuple[str, str], Tuple[int, float]]:
        """Return {(project_name, task_name): (soi_level, planned_pct)} for all projects."""
        cur = self.conn.cursor()
        cur.execute(
            """
            SELECT p.name, sp.task_name, sp.soi_level, sp.planned_pct
            FROM project_soi_plan sp
            JOIN projects p ON p.id = sp.project_id
            """
        )
        out: Dict[Tuple[str, str], Tuple[int, float]] = {}
        for pname, task, soi, pct in cur.fetchall():
            out[(str(pname), str(task))] = (int(soi or 1), float(pct or 0.0))
        return out

    def delete_project(self, project_key: str) -> None:
        """Delete a project and all sessions tied to it.

        Accepts:
          - project id (uuid or 'p1' style), OR
          - project code (e.g., 'PROJ-001'), OR
          - UI label (e.g., 'PROJ-001 — E-Commerce Website')
        """
        project_key = (project_key or "").strip()
        if not project_key:
            return

        # If the UI passed a label, extract the code.
        if "—" in project_key:
            project_key = project_key.split("—", 1)[0].strip()

        cur = self.conn.cursor()

        # Resolve project id.
        pid = None
        cur.execute("SELECT id FROM projects WHERE id = ?", (project_key,))
        row = cur.fetchone()
        if row:
            pid = row[0]
        else:
            cur.execute("SELECT id FROM projects WHERE UPPER(code)=UPPER(?)", (project_key,))
            row = cur.fetchone()
            if row:
                pid = row[0]

        if not pid:
            # Nothing to delete
            return

        cur.execute("DELETE FROM sessions WHERE project_id=?", (pid,))
        cur.execute("DELETE FROM projects WHERE id=?", (pid,))
        self.conn.commit()

    def reset_demo(self) -> None:
        cur = self.conn.cursor()
        cur.execute("DELETE FROM sessions")
        cur.execute("DELETE FROM projects")
        cur.execute("DELETE FROM employees")
        cur.execute("DELETE FROM users")
        self.conn.commit()
        self.seed_demo()

    def load(self) -> AppData:
        d = AppData()
        cur = self.conn.cursor()

        cur.execute("SELECT id,name,role FROM users ORDER BY role DESC, name ASC")
        d.users = [User(*row) for row in cur.fetchall()]

        cur.execute(
            "SELECT id, name, COALESCE(emp_code,''), COALESCE(laptop_brand,''), COALESCE(laptop_no,''), COALESCE(email,'') "
            "FROM employees ORDER BY name ASC"
        )
        d.employees = [
            Employee(id=row[0], name=row[1], emp_code=row[2], laptop_brand=row[3], laptop_no=row[4], email=row[5] if len(row)>5 else "")
            for row in cur.fetchall()
        ]

        cur.execute("SELECT id,code,name,use_module,allowed_tasks,COALESCE(planned_hours,0) FROM projects ORDER BY code ASC")
        d.projects = [Project(id=row[0], code=row[1], name=row[2], use_module=row[3], allowed_tasks=row[4], planned_hours=float(row[5] or 0.0)) for row in cur.fetchall()]

        cur.execute("SELECT id,employee_id,project_id,module,task_category,remark,punch_in,punch_out FROM sessions ORDER BY punch_in ASC")
        d.sessions = [
            WorkSession(
                id=row[0],
                employee_id=row[1],
                project_id=row[2],
                module=normalize_module(row[3]),
                task_category=row[4],
                remark=row[5] or "",
                punch_in=iso_to_dt(row[6]) or now(),
                punch_out=iso_to_dt(row[7]),
            )
            for row in cur.fetchall()
        ]
        return d

    def upsert_session(self, s: WorkSession) -> None:
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO sessions(id,employee_id,project_id,module,task_category,remark,punch_in,punch_out)
            VALUES (?,?,?,?,?,?,?,?)
            ON CONFLICT(id) DO UPDATE SET
                employee_id=excluded.employee_id,
                project_id=excluded.project_id,
                module=excluded.module,
                task_category=excluded.task_category,
                remark=excluded.remark,
                punch_in=excluded.punch_in,
                punch_out=excluded.punch_out
            """,
            (
                s.id,
                s.employee_id,
                s.project_id,
                (normalize_module(s.module) or s.module),
                s.task_category,
                s.remark,
                dt_to_iso(s.punch_in),
                dt_to_iso(s.punch_out),
            ),
        )
        self.conn.commit()

    def delete_session(self, session_id: str) -> None:
        """Delete a work session by id."""
        cur = self.conn.cursor()
        cur.execute("DELETE FROM sessions WHERE id=?", (session_id,))
        self.conn.commit()


    
    # -----------------------------
    # Approvals (admin gating export)
    # -----------------------------
    def set_approval(self, employee_id: str, d_from: date, d_to: date, approved: bool, approved_by: str = "ADMIN") -> None:
        employee_id = (employee_id or "").strip()
        if not employee_id:
            return
        df = fmt_date(d_from)
        dt = fmt_date(d_to)
        cur = self.conn.cursor()
        cur.execute(
            """
            INSERT INTO approvals(employee_id, date_from, date_to, approved, approved_at, approved_by)
            VALUES (?,?,?,?,?,?)
            ON CONFLICT(employee_id, date_from, date_to) DO UPDATE SET
                approved=excluded.approved,
                approved_at=excluded.approved_at,
                approved_by=excluded.approved_by
            """,
            (employee_id, df, dt, 1 if approved else 0, dt_to_iso(now()) if approved else None, approved_by),
        )
        self.conn.commit()

    def is_approved(self, employee_id: str, d_from: date, d_to: date) -> bool:
        employee_id = (employee_id or "").strip()
        if not employee_id:
            return False
        df = fmt_date(d_from)
        dt = fmt_date(d_to)
        cur = self.conn.cursor()
        cur.execute(
            "SELECT approved FROM approvals WHERE employee_id=? AND date_from=? AND date_to=?",
            (employee_id, df, dt),
        )
        row = cur.fetchone()
        return bool(row and int(row[0]) == 1)

    def approved_employee_ids(self, d_from: date, d_to: date) -> List[str]:
        df = fmt_date(d_from)
        dt = fmt_date(d_to)
        cur = self.conn.cursor()
        cur.execute(
            "SELECT employee_id FROM approvals WHERE date_from=? AND date_to=? AND approved=1",
            (df, dt),
        )
        return [r[0] for r in cur.fetchall()]

    def export_weekly_excel(
        self,
        out_path: str,
        d_from: Optional[date] = None,
        d_to: Optional[date] = None,
        *,
        # Backward compatible aliases (some App.py versions call with d1/d2)
        d1: Optional[date] = None,
        d2: Optional[date] = None,
        # Ignored compatibility args (older UI passes these)
        role: Optional[str] = None,
        employee_ids: Optional[set[str]] = None,
        break_seconds_per_day: int = 1800,
    ) -> None:
        """Export weekly report to Excel (wrapper used by App.py).

        - d_from / d_to are inclusive date bounds.
        - If employee_ids is provided, exports only those employees.
        """
        # Resolve date range
        if d_from is None and d1 is not None:
            d_from = d1
        if d_to is None and d2 is not None:
            d_to = d2

        if d_from is None or d_to is None:
            raise ValueError("export_weekly_excel requires a date range (d_from/d_to)")

        data = self.load()

        # ✅ Export only APPROVED employees for the selected date range
        approved_ids = set(self.approved_employee_ids(d_from, d_to))
        if employee_ids is None:
            employee_ids = approved_ids
        else:
            employee_ids = set(employee_ids) & approved_ids

        if employee_ids is not None:
            data = AppData(
                users=data.users,
                employees=data.employees,
                projects=data.projects,
                sessions=[s for s in data.sessions if s.employee_id in employee_ids],
            )
        export_summaries_only_xlsx(data, out_path, d_from, d_to, store=self)
        return
    def close(self) -> None:
        try:
            self.conn.close()
        except Exception:
            pass
# =============================
# Theme (professional light)
# =============================
class Theme:
    BG = "#f5f7fb"
    SURFACE = "#ffffff"
    TEXT = "#0f172a"
    MUTED = "#64748b"

    PRIMARY = "#2b6cb0"
    PRIMARY_DARK = "#1f5a96"
    PRIMARY_SOFT = "#eef5ff"
    PRIMARY_SOFT_2 = "#dbeafe"

    BORDER = "#d9e2ef"
    GRID = "#e9eff7"

    WARN = "#b45309"



def export_summaries_only_xlsx(data: "AppData", out_path: str, d_from: date, d_to: date, store=None) -> None:
    """
    Export ONLY these 3 sheets (selected date range):
      1) Project Summary
      2) Project + Module
      3) Employee Specific

    This is what the UI Summary tabs show, and nothing else.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # remove default sheet
    try:
        wb.remove(wb.active)
    except Exception:
        pass

    # name maps
    proj_name = {p.id: (p.name or p.code or p.id) for p in data.projects}
    proj_planned = { (p.name or p.code or p.id): float(getattr(p, 'planned_hours', 0.0) or 0.0) for p in data.projects }
    emp_name = {e.id: (e.name or e.id) for e in data.employees}

    def _iter_closed_sessions():
        for s in data.sessions:
            if getattr(s, "is_open", False) or not getattr(s, "punch_out", None):
                continue
            sd = s.punch_in.date()
            if sd < d_from or sd > d_to:
                continue
            yield s

    # -----------------------------
    # Sheet 1: Project Summary
    # -----------------------------
    ws4 = wb.create_sheet("Project Summary")
    ws4["A1"] = "PROJECT SUMMARY (SELECTED DATES)"
    ws4["A1"].font = Font(bold=True, size=14, color="111827")
    ws4.merge_cells("A1:D1")
    ws4["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws4.merge_cells("A2:D2")

    ws4_headers = ["Project", "Planned Hours", "Used Hours", "Remaining"]
    for c, h in enumerate(ws4_headers, start=1):
        ws4.cell(row=4, column=c, value=h)

    proj_used = {}
    for s in _iter_closed_sessions():
        pname = proj_name.get(s.project_id, s.project_id)
        proj_used[pname] = proj_used.get(pname, 0.0) + (s.duration_seconds() / 3600.0)

    r = 5
    all_projects = sorted(set(list(proj_planned.keys()) + list(proj_used.keys())), key=lambda x: x.lower())
    for pname in all_projects:
        planned = float(proj_planned.get(pname, 0.0) or 0.0)
        used = float(proj_used.get(pname, 0.0) or 0.0)
        remaining = planned - used if planned > 0 else 0.0
        ws4.cell(row=r, column=1, value=pname)
        ws4.cell(row=r, column=2, value=round(planned, 2))
        ws4.cell(row=r, column=3, value=round(used, 2))
        ws4.cell(row=r, column=4, value=round(remaining, 2))
        r += 1

    if r == 5:
        ws4.cell(row=5, column=1, value="No data")
        ws4.merge_cells("A5:D5")
        r = 6

    _apply_table_style(ws4, header_row=4, start_col=1, end_col=4, start_row=4, end_row=r-1)
    for col, width in enumerate([28, 14, 14, 14], start=1):
        ws4.column_dimensions[get_column_letter(col)].width = width

    # -----------------------------
    # Sheet 2: Project + Module
    # -----------------------------
    ws5 = wb.create_sheet("Project + Module")
    ws5["A1"] = "PROJECT + MODULE SUMMARY (SELECTED DATES)"
    ws5["A1"].font = Font(bold=True, size=14, color="111827")
    ws5.merge_cells("A1:C1")
    ws5["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws5.merge_cells("A2:C2")

    ws5_headers = ["Project", "Module", "Total Hours"]
    for c, h in enumerate(ws5_headers, start=1):
        ws5.cell(row=4, column=c, value=h)

    pm_totals = {}
    for s in _iter_closed_sessions():
        pname = proj_name.get(s.project_id, s.project_id)
        mod = getattr(s, "module", "") or "-"
        pm_totals[(pname, mod)] = pm_totals.get((pname, mod), 0.0) + (s.duration_seconds() / 3600.0)

    r = 5
    for (pname, mod), hrs in sorted(pm_totals.items(), key=lambda x: (x[0][0].lower(), str(x[0][1]).lower())):
        ws5.cell(row=r, column=1, value=pname)
        ws5.cell(row=r, column=2, value=mod)
        ws5.cell(row=r, column=3, value=round(float(hrs), 2))
        r += 1

    if r == 5:
        ws5.cell(row=5, column=1, value="No data")
        ws5.merge_cells("A5:C5")
        r = 6

    _apply_table_style(ws5, header_row=4, start_col=1, end_col=3, start_row=4, end_row=r-1)
    for col, width in enumerate([28, 18, 14], start=1):
        ws5.column_dimensions[get_column_letter(col)].width = width

    # -----------------------------
    # Sheet 3: Employee Specific
    # -----------------------------
    ws6 = wb.create_sheet("Employee Specific")
    ws6["A1"] = "EMPLOYEE SPECIFIC SUMMARY (SELECTED DATES)"
    ws6["A1"].font = Font(bold=True, size=14, color="111827")
    ws6.merge_cells("A1:F1")
    ws6["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws6.merge_cells("A2:F2")

    ws6_headers = ["Employee", "Project", "Module", "Status", "Hours", "Approved"]
    for c, h in enumerate(ws6_headers, start=1):
        ws6.cell(row=4, column=c, value=h)
    # Build totals by employee/project/module/status (status=Closed always for selected closed sessions)
    totals = {}
    for s in _iter_closed_sessions():
        en = emp_name.get(s.employee_id, s.employee_id)
        pname = proj_name.get(s.project_id, s.project_id)
        mod = getattr(s, "module", "") or "-"
        key = (en, pname, mod)
        totals[key] = totals.get(key, 0.0) + (s.duration_seconds() / 3600.0)

    r = 5
    # group by employee then project then module
    for (en, pname, mod), hrs in sorted(totals.items(), key=lambda x: (x[0][0].lower(), x[0][1].lower(), str(x[0][2]).lower())):
        ws6.cell(row=r, column=1, value=en)
        ws6.cell(row=r, column=2, value=pname)
        ws6.cell(row=r, column=3, value=mod)
        ws6.cell(row=r, column=4, value="Closed")
        ws6.cell(row=r, column=5, value=round(float(hrs), 2))
        ws6.cell(row=r, column=6, value="")  # approvals filled by DataStore wrapper if needed
        r += 1

    if r == 5:
        ws6.cell(row=5, column=1, value="No data")
        ws6.merge_cells("A5:F5")
        r = 6

    
    # Merge repeated Employee names in column A for better readability
    try:
        data_start = 5
        data_end = r - 1
        if data_end >= data_start:
            cur_val = None
            grp_start = data_start
            for rr in range(data_start, data_end + 2):  # +1 sentinel
                val = ws6.cell(row=rr, column=1).value if rr <= data_end else None
                if rr == data_start:
                    cur_val = val
                    grp_start = data_start
                    continue
                if val != cur_val:
                    grp_end = rr - 1
                    if cur_val not in (None, "") and grp_end > grp_start:
                        ws6.merge_cells(start_row=grp_start, start_column=1, end_row=grp_end, end_column=1)
                        top = ws6.cell(row=grp_start, column=1)
                        top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cur_val = val
                    grp_start = rr
    except Exception:
        pass

    _apply_table_style(ws6, header_row=4, start_col=1, end_col=6, start_row=4, end_row=r-1)
    for col, width in enumerate([20, 28, 16, 12, 12, 12], start=1):
            ws6.column_dimensions[get_column_letter(col)].width = width

            
    # -----------------------------
    # Sheet 4: SOI Summary
    # -----------------------------
    ws7 = wb.create_sheet("SOI Summary")
    ws7["A1"] = "SOI SUMMARY (SELECTED DATES)"
    ws7["A1"].font = Font(bold=True, size=14, color="111827")
    ws7.merge_cells("A1:G1")
    ws7["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws7.merge_cells("A2:G2")

    ws7_headers = ["Project", "Module", "Task", "SOI", "Planned Hours", "Used Hours", "Diff (Planned-Used)"]
    for c, h in enumerate(ws7_headers, start=1):
        ws7.cell(row=4, column=c, value=h)

    # Build SOI plan map (task->soi) and compute planned HOURS per module using planned_pct.
    # planned_pct is a percentage of module planned hours (or project planned hours if module plan missing).
    soi_map = {}    # (project_id, task) -> soi_label
    soi_pct = {}    # (project_id, task, soi_label) -> planned_pct
    if store is not None:
        try:
            for p in data.projects:
                rows = store.get_project_soi_plan(p.id)  # [(task_name, soi_level, planned_pct), ...]
                for task, soi_level, planned_pct in rows:
                    if not task:
                        continue
                    soi_label = f"SOI-{int(soi_level or 1)}"
                    soi_map[(p.id, task)] = soi_label
                    soi_pct[(p.id, task, soi_label)] = float(planned_pct or 0.0)
        except Exception:
            pass

    # Planned hours by (project, module, task, soi)
    planned = {}  # (project_name, module_name, task, soi_label) -> planned_hours
    if store is not None:
        try:
            for p in data.projects:
                pname = proj_name.get(p.id, p.id)

                # module planned hours
                modules = []
                if hasattr(store, "get_project_modules"):
                    modules = store.get_project_modules(p.id)  # [(module_name, planned_hours), ...]

                if modules:
                    module_rows = [(m, float(ph or 0.0)) for (m, ph) in modules]
                else:
                    # fallback: use overall project planned hours if module plan not configured
                    module_rows = [("-", float(getattr(p, "planned_hours", 0.0) or 0.0))]

                # apply SOI planned percentages on each module planned hours
                for (mod_name, mod_planned_hrs) in module_rows:
                    for (pid, task, soi_label), pct in list(soi_pct.items()):
                        if pid != p.id:
                            continue
                        hrs = (mod_planned_hrs * float(pct)) / 100.0 if mod_planned_hrs and pct else 0.0
                        key = (pname, mod_name or "-", task, soi_label)
                        planned[key] = planned.get(key, 0.0) + hrs
        except Exception:
            pass

    # Used hours by (project, module, task, soi)
    used = {}
    for s in _iter_closed_sessions():
        pname = proj_name.get(s.project_id, s.project_id)
        mod = getattr(s, "module", "") or "-"
        task = getattr(s, "task_category", "") or "-"
        soi_label = soi_map.get((s.project_id, task), "-")
        key = (pname, mod, task, soi_label)
        used[key] = used.get(key, 0.0) + (s.duration_seconds() / 3600.0)

    all_keys = sorted(set(list(planned.keys()) + list(used.keys())),
                      key=lambda k: (k[0].lower(), str(k[1]).lower(), str(k[3]).lower(), str(k[2]).lower()))

    r7 = 5
    for (pname, mod, task, soi_label) in all_keys:
        ph = float(planned.get((pname, mod, task, soi_label), 0.0) or 0.0)
        uh = float(used.get((pname, mod, task, soi_label), 0.0) or 0.0)
        ws7.cell(row=r7, column=1, value=pname)
        ws7.cell(row=r7, column=2, value=mod or "-")
        ws7.cell(row=r7, column=3, value=task or "-")
        ws7.cell(row=r7, column=4, value=soi_label or "-")
        ws7.cell(row=r7, column=5, value=round(ph, 2))
        ws7.cell(row=r7, column=6, value=round(uh, 2))
        ws7.cell(row=r7, column=7, value=round(ph - uh, 2))
        r7 += 1

    if r7 == 5:
        ws7.cell(row=5, column=1, value="No SOI data")
        ws7.merge_cells("A5:G5")
        r7 = 6

        # Merge repeated Project (col A) and Module (col B) cells for readability
    try:
        data_start = 5
        data_end = r7 - 1
        if data_end >= data_start:
            # Merge Project column (A) for consecutive identical project names
            cur_proj = None
            grp_start = data_start
            for rr in range(data_start, data_end + 2):  # sentinel
                val = ws7.cell(row=rr, column=1).value if rr <= data_end else None
                if rr == data_start:
                    cur_proj = val
                    grp_start = data_start
                    continue
                if val != cur_proj:
                    grp_end = rr - 1
                    if cur_proj not in (None, "") and grp_end > grp_start:
                        ws7.merge_cells(start_row=grp_start, start_column=1, end_row=grp_end, end_column=1)
                        top = ws7.cell(row=grp_start, column=1)
                        top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cur_proj = val
                    grp_start = rr

            # Merge Module column (B) within each Project group
            rr = data_start
            while rr <= data_end:
                proj_val = ws7.cell(row=rr, column=1).value
                # find end of this project group
                rr_end = rr
                while rr_end + 1 <= data_end and ws7.cell(row=rr_end + 1, column=1).value == proj_val:
                    rr_end += 1

                # merge module inside [rr, rr_end]
                cur_mod = ws7.cell(row=rr, column=2).value
                mod_start = rr
                for rrr in range(rr, rr_end + 2):  # sentinel
                    mod_val = ws7.cell(row=rrr, column=2).value if rrr <= rr_end else None
                    if rrr == rr:
                        cur_mod = mod_val
                        mod_start = rr
                        continue
                    if mod_val != cur_mod:
                        mod_end = rrr - 1
                        if cur_mod not in (None, "") and mod_end > mod_start:
                            ws7.merge_cells(start_row=mod_start, start_column=2, end_row=mod_end, end_column=2)
                            top = ws7.cell(row=mod_start, column=2)
                            top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        cur_mod = mod_val
                        mod_start = rrr

                rr = rr_end + 1
    except Exception:
        pass

    _apply_table_style(ws7, header_row=4, start_col=1, end_col=7, start_row=4, end_row=r7-1)
    for col, width in enumerate([28, 14, 22, 10, 14, 14, 18], start=1):
        ws7.column_dimensions[get_column_letter(col)].width = width


    wb.save(out_path)


    # =============================
    # Excel export helpers
    # =============================
def _apply_table_style(ws, header_row: int, start_col: int, end_col: int, last_row: int = None, *,
                        start_row: int = None, end_row: int = None,
                        header_fill="CFE8F7", header_font_color="0B3D91",
                        zebra_fill="FFF7CC", body_fill=None,
                        border_color="1F2937", border_style="medium",
                        emphasize_last_col=False, last_col_fill="F4B183",
                        freeze=True):
        """
        Applies a professional, high-contrast table style similar to your sample:
        - Light blue header
        - Yellow body/zebra
        - Dark borders
        - Optional emphasized last column (orange)
        """
        hdr_fill = PatternFill("solid", fgColor=header_fill)
        hdr_font = Font(bold=True, color=header_font_color)
        body_font = Font(color="111827")
        zebra = PatternFill("solid", fgColor=zebra_fill) if zebra_fill else None
        body = PatternFill("solid", fgColor=body_fill) if body_fill else None

        side = Side(style=border_style, color=border_color)
        border = Border(left=side, right=side, top=side, bottom=side)

        # Backward/forward compatible row range
        table_start = start_row if start_row is not None else header_row
        table_end = end_row if end_row is not None else (last_row if last_row is not None else header_row)
        if table_end < header_row:
            table_end = header_row

        # Header
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.border = border

        # Body
        for r in range(header_row + 1, table_end + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if body is not None:
                    cell.fill = body
            if zebra and (r - header_row) % 2 == 0:
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).fill = zebra

        # Optional emphasize last column (like orange % bar in sample)
        if emphasize_last_col:
            fill_last = PatternFill("solid", fgColor=last_col_fill)
            for r in range(header_row, table_end + 1):
                ws.cell(row=r, column=end_col).fill = fill_last
                ws.cell(row=r, column=end_col).border = border
                ws.cell(row=r, column=end_col).alignment = Alignment(vertical="center", horizontal="center")

        if freeze:
            ws.freeze_panes = ws.cell(row=header_row + 1, column=start_col)

        # Autosize columns
        for c in range(start_col, end_col + 1):
            max_len = 0
            col_letter = get_column_letter(c)
            for r in range(1, table_end + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 52)

    # Compatibility alias (older app versions)
        _apply_table_style0 = _apply_table_style



        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=header_row, column=c)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(vertical="center", horizontal="center")
            cell.border = border

        for r in range(header_row + 1, table_end + 1):
            for c in range(start_col, end_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = body_font
                cell.border = border
                cell.alignment = Alignment(vertical="center")
            if (r - header_row) % 2 == 0:
                for c in range(start_col, end_col + 1):
                    ws.cell(row=r, column=c).fill = zebra

        ws.freeze_panes = ws.cell(row=header_row + 1, column=start_col)

        for c in range(start_col, end_col + 1):
            max_len = 0
            col_letter = get_column_letter(c)
            for r in range(header_row, table_end + 1):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 44)


def export_weekly_report_xlsx(data: AppData, out_path: str, d_from: date, d_to: date, break_seconds_per_day: int = 1800):
    """
    Excel export (professional colors + dark borders).

    Sheet 1: Employee → Project → Task → day-wise table (Mon..Sun of the week of d_from)
      - If selected range spans beyond that week, extra days are also included.
      - Adds EMPLOYEE TOTAL row (day-wise totals) for each employee.
      - Adds GRAND TOTAL row (day-wise totals).

    Sheet 2: Project + Task Summary for SELECTED dates ONLY (Total, not week total).

    Sheet 3: Employee + Project + Task totals for SELECTED dates ONLY.

    Break rule:
      - For each employee/day, subtract 1.5 hours if any work logged that day.
      - The deduction is distributed proportionally across that day's (project, task) buckets.
    """
    def _bucket_parts(b):
        """Return (pid, module, task) from legacy/new bucket tuples."""
        try:
            pid = b[0] if len(b) > 0 else ""
            module = b[1] if len(b) > 1 else ""
            task = b[2] if len(b) > 2 else ""
            return pid, module, task
        except Exception:
            return "", "", ""

    # Display (Sheet-1) covers at least Mon..Sun of week containing d_from
    display_start = start_of_week(d_from)  # Monday
    display_end = max(d_to, display_start + timedelta(days=6))  # at least Sunday
    display_days = daterange(display_start, display_end)

    selected_days = daterange(d_from, d_to)
    selected_set = set(selected_days)

    range_start = datetime.combine(display_start, time(0, 0, 0))
    range_end = datetime.combine(display_end + timedelta(days=1), time(0, 0, 0))

    # employee -> day -> (project_id, task) -> seconds (raw)
    raw: Dict[str, Dict[date, Dict[Tuple[str, str, str], int]]] = {}

    for s in data.sessions:
        if s.is_open or not s.punch_out:
            continue
        if s.punch_out <= range_start or s.punch_in >= range_end:
            continue

        start_c = clamp_dt(s.punch_in, range_start, range_end)
        end_c = clamp_dt(s.punch_out, range_start, range_end)
        if end_c <= start_c:
            continue

        per_day = split_session_by_day(start_c, end_c)
        bucket = (s.project_id, s.module, s.task_category)
        for dday, sec in per_day.items():
            raw.setdefault(s.employee_id, {}).setdefault(dday, {}).setdefault(bucket, 0)
            raw[s.employee_id][dday][bucket] += sec

    # Apply break: employee -> day -> bucket -> seconds (net)
    net: Dict[str, Dict[date, Dict[Tuple[str, str, str], int]]] = {}
    for emp_id, by_day in raw.items():
        for dday, by_bucket in by_day.items():
            total_day = sum(by_bucket.values())
            if total_day <= 0:
                continue

            break_to_apply = min(break_seconds_per_day, total_day)
            net.setdefault(emp_id, {}).setdefault(dday, {})

            for bucket, sec in by_bucket.items():
                share = (sec / total_day) if total_day else 0.0
                deduct = int(round(break_to_apply * share))
                net[emp_id][dday][bucket] = max(0, sec - deduct)

            # rounding fix to match target = total_day - break
            target = total_day - break_to_apply
            current = sum(net[emp_id][dday].values())
            diff = target - current
            if diff != 0 and net[emp_id][dday]:
                largest_bucket = max(net[emp_id][dday].items(), key=lambda kv: kv[1])[0]
                net[emp_id][dday][largest_bucket] = max(0, net[emp_id][dday][largest_bucket] + diff)

    emp_name = {e.id: e.name for e in data.employees}
    proj_name = {p.id: (p.name or p.code or p.id) for p in data.projects}

    # -----------------------------
    # Sheet 1
    # -----------------------------
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Employee Project Task"

    total_cols = 4 + len(display_days) + 1

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws1["A1"] = "WEEKLY TIME REPORT (NET OF DAILY BREAK)"
    ws1["A1"].font = Font(bold=True, size=14, color="111827")
    ws1["A1"].fill = PatternFill("solid", fgColor="BFBFBF")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws1["A2"] = f"Date range: {fmt_date(display_start)} to {fmt_date(display_end)}   |   Break: 01:30 per employee/day (if any work logged)"
    ws1["A2"].font = Font(color="111827")
    ws1["A2"].fill = PatternFill("solid", fgColor="E5E7EB")
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")

    header_row = 4
    headers = ["Employee", "Project", "Module", "Task"] + [d.strftime("%a %d-%b") for d in display_days] + ["Total (HH:MM)"]
    for i, h in enumerate(headers, start=1):
        ws1.cell(row=header_row, column=i, value=h)

    r = header_row + 1

    employee_row_spans: List[Tuple[int, int]] = []
    employee_total_rows: List[int] = []

    grand_by_day = {d: 0 for d in display_days}
    grand_total = 0

    for emp_id in sorted(emp_name.keys(), key=lambda x: emp_name.get(x, "").lower()):
        # collect all (project,task) buckets across display days
        buckets = set()
        for dday in display_days:
            buckets |= set(net.get(emp_id, {}).get(dday, {}).keys())
        if not buckets:
            continue

        emp_start = r
        buckets_sorted = sorted(list(buckets), key=lambda b: (proj_name.get(_bucket_parts(b)[0], _bucket_parts(b)[0]), _bucket_parts(b)[1], _bucket_parts(b)[2]))

        emp_by_day = {d: 0 for d in display_days}
        emp_total = 0

        for b in buckets_sorted:
            pid, module, task = _bucket_parts(b)
            total_bucket = 0
            day_secs = []
            for dday in display_days:
                sec = net.get(emp_id, {}).get(dday, {}).get((pid, module, task), 0)
                day_secs.append(sec)
                total_bucket += sec
                emp_by_day[dday] += sec
                grand_by_day[dday] += sec

            emp_total += total_bucket
            grand_total += total_bucket

            row = [emp_name[emp_id], proj_name.get(pid, pid), module, task] + [seconds_to_hhmm(s) for s in day_secs] + [seconds_to_hhmm(total_bucket)]
            for c, val in enumerate(row, start=1):
                ws1.cell(row=r, column=c, value=val)
            r += 1

        # Employee TOTAL row (day-wise)
        emp_total_row = r
        employee_total_rows.append(emp_total_row)
        row = [emp_name[emp_id], "EMPLOYEE TOTAL", "", ""] + [seconds_to_hhmm(emp_by_day[d]) for d in display_days] + [seconds_to_hhmm(emp_total)]
        for c, val in enumerate(row, start=1):
            ws1.cell(row=emp_total_row, column=c, value=val)
        r += 1

        emp_end = r - 1
        employee_row_spans.append((emp_start, emp_end))

        # separator row
        sep_fill = PatternFill("solid", fgColor="E6F0FA")
        for c in range(1, len(headers) + 1):
            ws1.cell(row=r, column=c, value=None).fill = sep_fill
        r += 1

    # GRAND TOTAL row
    grand_row = r
    row = ["GRAND TOTAL", "", "", ""] + [seconds_to_hhmm(grand_by_day[d]) for d in display_days] + [seconds_to_hhmm(grand_total)]
    for c, val in enumerate(row, start=1):
        ws1.cell(row=grand_row, column=c, value=val)

    last_row = grand_row

    _apply_table_style(
        ws1,
        header_row=header_row,
        start_col=1,
        end_col=len(headers),
        last_row=last_row,
        header_fill="CFE8F7",
        zebra_fill="FFF7CC",
        border_color="111827",
        border_style="medium",
        emphasize_last_col=True,
        last_col_fill="F4B183",
    )

    # Column fills
    emp_fill = PatternFill("solid", fgColor="FFF2CC")
    proj_fill = PatternFill("solid", fgColor="FFF7CC")
    task_fill = PatternFill("solid", fgColor="E0F2FE")
    emp_font = Font(bold=True, color="111827")
    emp_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # Merge Employee cells per block (including total row)
    for (rs, re_) in employee_row_spans:
        if re_ > rs:
            ws1.merge_cells(start_row=rs, start_column=1, end_row=re_, end_column=1)
        c0 = ws1.cell(row=rs, column=1)
        c0.fill = emp_fill
        c0.font = emp_font
        c0.alignment = emp_align
        for rr in range(rs, re_ + 1):
            ws1.cell(row=rr, column=1).fill = emp_fill

    for rr in range(header_row + 1, table_end + 1):
        ws1.cell(row=rr, column=2).fill = proj_fill
        ws1.cell(row=rr, column=3).fill = task_fill

    # Highlight employee total rows
    emp_total_fill = PatternFill("solid", fgColor="D9F2FF")
    for rr in employee_total_rows:
        for c in range(1, len(headers) + 1):
            ws1.cell(row=rr, column=c).fill = emp_total_fill
            ws1.cell(row=rr, column=c).font = Font(bold=True, color="111827")
            ws1.cell(row=rr, column=c).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=rr, column=len(headers)).fill = PatternFill("solid", fgColor="F4B183")

    # Grand total row green
    green_fill = PatternFill("solid", fgColor="A7F3A0")
    for c in range(1, len(headers) + 1):
        ws1.cell(row=grand_row, column=c).fill = green_fill
        ws1.cell(row=grand_row, column=c).font = Font(bold=True, color="111827")
        ws1.cell(row=grand_row, column=c).alignment = Alignment(horizontal="center", vertical="center")

    ws1.row_dimensions[1].height = 26
    ws1.row_dimensions[2].height = 20

    # -----------------------------
    # Sheet 2: Project + Task summary (selected dates)
    # -----------------------------
    ws2 = wb.create_sheet("Project Task Summary")

    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws2["A1"] = "PROJECT + TASK SUMMARY (SELECTED DATES)"
    ws2["A1"].font = Font(bold=True, size=14, color="111827")
    ws2["A1"].fill = PatternFill("solid", fgColor="BFBFBF")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws2["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws2["A2"].font = Font(color="111827")
    ws2["A2"].fill = PatternFill("solid", fgColor="E5E7EB")
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")

    header_row2 = 4
    headers2 = ["Project", "Task", "Module", "Total (HH:MM)", "Share %"]
    for i, h in enumerate(headers2, start=1):
        ws2.cell(row=header_row2, column=i, value=h)

    bucket_totals: Dict[Tuple[str, str, str], int] = {}
    for emp_id, by_day in net.items():
        for dday, by_bucket in by_day.items():
            if dday not in selected_set:
                continue
            for (pid, module, task), sec in by_bucket.items():
                key = (pid, module, task)
                bucket_totals[key] = bucket_totals.get(key, 0) + sec

    company_total = sum(bucket_totals.values()) or 1

    r2 = header_row2 + 1
    for b in sorted(bucket_totals.keys(), key=lambda x: (proj_name.get(_bucket_parts(x)[0], _bucket_parts(x)[0]), _bucket_parts(x)[1], _bucket_parts(x)[2])):
        pid, module, task = _bucket_parts(b)
        sec = bucket_totals[b]
        ws2.cell(row=r2, column=1, value=proj_name.get(pid, pid))
        ws2.cell(row=r2, column=2, value=task)
        ws2.cell(row=r2, column=3, value=module)
        ws2.cell(row=r2, column=4, value=seconds_to_hhmm(sec))
        ws2.cell(row=r2, column=5, value=round((sec / company_total) * 100, 2))
        r2 += 1

    total_row2 = r2
    ws2.cell(row=total_row2, column=1, value="GRAND TOTAL")
    ws2.cell(row=total_row2, column=2, value="")
    ws2.cell(row=total_row2, column=3, value="")
    ws2.cell(row=total_row2, column=4, value=seconds_to_hhmm(company_total))
    ws2.cell(row=total_row2, column=5, value=100.0)

    _apply_table_style(
        ws2,
        header_row=header_row2,
        start_col=1,
        end_col=len(headers2),
        last_row=total_row2,
        header_fill="CFE8F7",
        zebra_fill="FFF7CC",
        border_color="111827",
        border_style="medium",
        emphasize_last_col=True,
        last_col_fill="9DECF9",
    )

    for rr in range(header_row2 + 1, total_row2 + 1):
        ws2.cell(row=rr, column=5).number_format = "0.00"
        ws2.cell(row=rr, column=5).alignment = Alignment(horizontal="center", vertical="center")

    for c in range(1, len(headers2) + 1):
        ws2.cell(row=total_row2, column=c).fill = green_fill
        ws2.cell(row=total_row2, column=c).font = Font(bold=True, color="111827")
        ws2.cell(row=total_row2, column=c).alignment = Alignment(horizontal="center", vertical="center")

    ws2.row_dimensions[1].height = 26
    ws2.row_dimensions[2].height = 20

    # -----------------------------
    # Sheet 3: Employee + Project + Task totals (selected dates)
    # -----------------------------
    ws3 = wb.create_sheet("Employee Project Task Total")

    ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws3["A1"] = "EMPLOYEE + PROJECT + TASK (SELECTED DATES)"
    ws3["A1"].font = Font(bold=True, size=14, color="111827")
    ws3["A1"].fill = PatternFill("solid", fgColor="BFBFBF")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws3.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws3["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws3["A2"].font = Font(color="111827")
    ws3["A2"].fill = PatternFill("solid", fgColor="E5E7EB")
    ws3["A2"].alignment = Alignment(horizontal="center", vertical="center")

    header_row3 = 4
    headers3 = ["Employee", "Project", "Module", "Task", "Total (HH:MM)"]
    for i, h in enumerate(headers3, start=1):
        ws3.cell(row=header_row3, column=i, value=h)

    emp_bucket_totals: Dict[Tuple[str, str, str, str], int] = {}
    for emp_id, by_day in net.items():
        for dday, by_bucket in by_day.items():
            if dday not in selected_set:
                continue
            for (pid, module, task), sec in by_bucket.items():
                key = (emp_id, pid, module, task)
                emp_bucket_totals[key] = emp_bucket_totals.get(key, 0) + sec

    r3 = header_row3 + 1
    grand3 = 0
    for (emp_id, pid, module, task) in sorted(emp_bucket_totals.keys(), key=lambda k: (emp_name.get(k[0], k[0]).lower(), proj_name.get(k[1], k[1]), k[2], k[3])):
        sec = emp_bucket_totals[(emp_id, pid, module, task)]
        grand3 += sec
        ws3.cell(row=r3, column=1, value=emp_name.get(emp_id, emp_id))
        ws3.cell(row=r3, column=2, value=proj_name.get(pid, pid))
        ws3.cell(row=r3, column=3, value=module)
        ws3.cell(row=r3, column=4, value=task)
        ws3.cell(row=r3, column=5, value=seconds_to_hhmm(sec))
        r3 += 1

    total_row3 = r3
    ws3.cell(row=total_row3, column=1, value="GRAND TOTAL")
    ws3.cell(row=total_row3, column=2, value="")
    ws3.cell(row=total_row3, column=3, value="")
    ws3.cell(row=total_row3, column=4, value="")
    ws3.cell(row=total_row3, column=5, value=seconds_to_hhmm(grand3))

    _apply_table_style(
        ws3,
        header_row=header_row3,
        start_col=1,
        end_col=len(headers3),
        last_row=total_row3,
        header_fill="CFE8F7",
        zebra_fill="FFF7CC",
        border_color="111827",
        border_style="medium",
        emphasize_last_col=True,
        last_col_fill="F4B183",
    )

    for c in range(1, len(headers3) + 1):
        ws3.cell(row=total_row3, column=c).fill = green_fill
        ws3.cell(row=total_row3, column=c).font = Font(bold=True, color="111827")
        ws3.cell(row=total_row3, column=c).alignment = Alignment(horizontal="center", vertical="center")

    ws3.row_dimensions[1].height = 26
    ws3.row_dimensions[2].height = 20

    
    # -----------------------------
    # Sheet 4: Project Summary (Selected Dates)
    # -----------------------------
    ws4 = wb.create_sheet("Project Summary")
    ws4["A1"] = "PROJECT SUMMARY (SELECTED DATES)"
    ws4["A1"].font = Font(bold=True, size=14, color="111827")
    ws4.merge_cells("A1:B1")
    ws4["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws4.merge_cells("A2:B2")

    ws4_headers = ["Project", "Total Hours"]
    for c, h in enumerate(ws4_headers, start=1):
        ws4.cell(row=4, column=c, value=h)

    proj_totals = {}
    for s in data.sessions:
        if getattr(s, "is_open", False) or not getattr(s, "punch_out", None):
            continue
        sd = s.punch_in.date()
        if sd < d_from or sd > d_to:
            continue
        pname = proj_name.get(s.project_id, s.project_id)
        proj_totals[pname] = proj_totals.get(pname, 0) + s.duration_seconds()

    r4 = 5
    for pname in sorted(proj_totals.keys(), key=lambda x: x.lower()):
        ws4.cell(row=r4, column=1, value=pname)
        ws4.cell(row=r4, column=2, value=seconds_to_hhmm(proj_totals[pname]))
        r4 += 1

    if r4 == 5:
        ws4.cell(row=5, column=1, value="No data")
        ws4.merge_cells("A5:B5")
        r4 = 6

    _apply_table_style(ws4, header_row=4, start_col=1, end_col=2, last_row=r4-1,
                       header_fill="CFE8F7", zebra_fill="FFF7CC", border_color="111827",
                       border_style="medium", emphasize_last_col=True, last_col_fill="F4B183")

    # ---- Sheet: Project + Module ----
    ws5 = wb.create_sheet("Project + Module")
    ws5["A1"] = "PROJECT + MODULE SUMMARY (SELECTED DATES)"
    ws5["A1"].font = Font(bold=True, size=14, color="111827")
    ws5.merge_cells("A1:C1")
    ws5["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws5.merge_cells("A2:C2")

    ws5_headers = ["Project", "Module", "Total Hours"]
    for c, h in enumerate(ws5_headers, start=1):
        ws5.cell(row=4, column=c, value=h)

    pm_totals = {}
    for s in data.sessions:
        if getattr(s, "is_open", False) or not getattr(s, "punch_out", None):
            continue
        sd = s.punch_in.date()
        if sd < d_from or sd > d_to:
            continue
        pname = proj_name.get(s.project_id, s.project_id)
        mod = getattr(s, "module", "") or "-"
        pm_totals[(pname, mod)] = pm_totals.get((pname, mod), 0) + s.duration_seconds()

    r5 = 5
    for (pname, mod) in sorted(pm_totals.keys(), key=lambda x: (x[0].lower(), x[1].lower())):
        ws5.cell(row=r5, column=1, value=pname)
        ws5.cell(row=r5, column=2, value=mod)
        ws5.cell(row=r5, column=3, value=seconds_to_hhmm(pm_totals[(pname, mod)]))
        r5 += 1

    if r5 == 5:
        ws5.cell(row=5, column=1, value="No data")
        ws5.merge_cells("A5:C5")
        r5 = 6

    _apply_table_style(ws5, header_row=4, start_col=1, end_col=3, last_row=r5-1,
                       header_fill="CFE8F7", zebra_fill="FFF7CC", border_color="111827",
                       border_style="medium", emphasize_last_col=True, last_col_fill="F4B183")

    # -----------------------------
    # Sheet 6: Employee Specific (Selected Dates)
    # -----------------------------
    ws6 = wb.create_sheet("Employee Specific")
    ws6["A1"] = "EMPLOYEE SPECIFIC SUMMARY (SELECTED DATES)"
    ws6["A1"].font = Font(bold=True, size=14, color="111827")
    ws6.merge_cells("A1:E1")
    ws6["A2"] = f"Selected date range: {fmt_date(d_from)} to {fmt_date(d_to)}"
    ws6.merge_cells("A2:E2")

    ws6_headers = ["Employee", "Project", "Module", "Total Hours", "Approved"]
    for c, h in enumerate(ws6_headers, start=1):
        ws6.cell(row=4, column=c, value=h)

    emp_totals = {}
    for s in data.sessions:
        if getattr(s, "is_open", False) or not getattr(s, "punch_out", None):
            continue
        sd = s.punch_in.date()
        if sd < d_from or sd > d_to:
            continue
        en = emp_name.get(s.employee_id, s.employee_id)
        pname = proj_name.get(s.project_id, s.project_id)
        mod = getattr(s, "module", "") or "-"
        emp_totals[(en, pname, mod)] = emp_totals.get((en, pname, mod), 0) + s.duration_seconds()

    r6 = 5
    for (en, pname, mod) in sorted(emp_totals.keys(), key=lambda x: (x[0].lower(), x[1].lower(), x[2].lower())):
        ws6.cell(row=r6, column=1, value=en)
        ws6.cell(row=r6, column=2, value=pname)
        ws6.cell(row=r6, column=3, value=mod)
        ws6.cell(row=r6, column=4, value=seconds_to_hhmm(emp_totals[(en, pname, mod)]))
        ws6.cell(row=r6, column=5, value="Approved")
        r6 += 1

    if r6 == 5:
        ws6.cell(row=5, column=1, value="No data")
        ws6.merge_cells("A5:E5")
        r6 = 6

    _apply_table_style(ws6, header_row=4, start_col=1, end_col=5, last_row=r6-1,
                       header_fill="CFE8F7", zebra_fill="FFF7CC", border_color="111827",
                       border_style="medium", emphasize_last_col=True, last_col_fill="F4B183")
    wb.save(out_path)