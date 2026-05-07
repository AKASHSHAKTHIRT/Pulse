"""
Time Tracker - Flask Web Backend (Thread-Safe)
===============================================
KEY FIX: SQLite connections cannot be shared across threads.
Each HTTP request gets its own DataStore (SQLite connection) via Flask's
per-request 'g' context. The connection is automatically closed when the
request ends via teardown_appcontext.
"""

from __future__ import annotations
import os
import tempfile
from datetime import date, timedelta
from functools import wraps

from flask import Flask, request, jsonify, session, send_file, g
from core import (
    DataStore, WorkSession, TASK_CATEGORIES, MODULE_CATEGORIES,
    now, fmt_dt, fmt_date, parse_date, parse_user_datetime, short_id,
    seconds_to_hhmmss, start_of_week, normalize_module,
)

# ── App setup ──────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder="static", static_url_path="/static")

# ── Configurable settings ───────────────────────────────────────
INACTIVITY_DAYS    = int(os.environ.get("INACTIVITY_DAYS",    "3"))
UNPUNCHED_HOURS    = int(os.environ.get("UNPUNCHED_HOURS",    "10"))
AUTO_CLOSE_HOUR    = int(os.environ.get("AUTO_CLOSE_HOUR",    "20"))
AUTO_CLOSE_ENABLED = os.environ.get("AUTO_CLOSE_ENABLED", "1") != "0"

app.secret_key = os.environ.get("SECRET_KEY", "timetracker-secret-2024")

DB_PATH        = os.environ.get("DB_PATH", "time_tracker.db")
ADMIN_ID       = "ADMIN"
ADMIN_PASSWORD = "Admin@123"

# ── Thread-safe DataStore ──────────────────────────────────────────────────────
#
# Flask handles requests in multiple threads simultaneously.
# SQLite connections are NOT thread-safe by default.
# Solution: use Flask g (request-scoped) to give each request its own connection.

def get_store() -> DataStore:
    if "store" not in g:
        g.store = DataStore(DB_PATH)
    return g.store

@app.teardown_appcontext
def _close_store(exc=None):
    s = g.pop("store", None)
    if s is not None:
        try: s.close()
        except: pass

# Seed demo data once at startup
def _seed_once():
    tmp = DataStore(DB_PATH)
    tmp.seed_demo()
    tmp.close()

_seed_once()

# ── Auth decorators ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return jsonify({"error": "Not authenticated"}), 401
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("role") != "admin":
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)
    return decorated

# ── Auth routes ────────────────────────────────────────────────────────────────

@app.route("/api/login", methods=["POST"])
def login():
    b   = request.json or {}
    lid = (b.get("login_id") or "").strip().upper()
    pw  = (b.get("password") or "").strip()
    if not lid or not pw:
        return jsonify({"error": "Login ID and password are required"}), 400
    if lid == ADMIN_ID:
        if pw == ADMIN_PASSWORD:
            session.update({"user": "ADMIN", "role": "admin", "emp_id": None, "emp_name": "Admin"})
            return jsonify({"role": "admin", "name": "Admin"})
        return jsonify({"error": "Invalid credentials"}), 401
    store = get_store()
    data  = store.load()
    emp   = data.employee_by_code(lid)
    if not emp:
        return jsonify({"error": "Employee not found"}), 401
    store.ensure_credential(lid)
    if not store.verify_password(lid, pw):
        return jsonify({"error": "Invalid credentials"}), 401
    alerts = store.auto_zero_unended_tasks(lid)
    session.update({"user": lid, "role": "employee", "emp_id": emp.id, "emp_name": emp.name})
    return jsonify({"role": "employee", "name": emp.name, "emp_id": emp.id,
                    "emp_code": emp.emp_code, "alerts": alerts})

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"ok": True})

@app.route("/api/me")
def me():
    if "user" not in session:
        return jsonify({"authenticated": False})
    return jsonify({"authenticated": True, "role": session.get("role"),
                    "name": session.get("emp_name"), "emp_id": session.get("emp_id")})

# ── Reference data ─────────────────────────────────────────────────────────────

@app.route("/api/data")
@login_required
def get_data():
    store = get_store()
    d = store.load()
    task_categories = store.get_task_categories()
    return jsonify({
        "employees": [{"id": e.id, "name": e.name, "emp_code": e.emp_code,
                       "laptop_brand": e.laptop_brand, "laptop_no": e.laptop_no}
                      for e in d.employees],
        "projects": [{"id": p.id, "code": p.code, "name": p.name,
                      "use_module": p.use_module, "allowed_tasks": p.allowed_tasks,
                      "planned_hours": p.planned_hours}
                     for p in d.projects],
        "task_categories": task_categories,
        "module_categories": MODULE_CATEGORIES,
    })

# ── Punch In / Out ─────────────────────────────────────────────────────────────

@app.route("/api/punch/status")
@login_required
def punch_status():
    eid = session.get("emp_id")
    if not eid: return jsonify({"error": "No employee in session"}), 400
    store = get_store(); d = store.load()
    os_ = d.get_open_session_for_employee(eid)
    if os_:
        proj = d.project_by_id(os_.project_id)
        return jsonify({"active": True, "session_id": os_.id,
                        "project": proj.name if proj else os_.project_id,
                        "project_id": os_.project_id, "module": os_.module,
                        "task_category": os_.task_category, "remark": os_.remark,
                        "punch_in": fmt_dt(os_.punch_in), "elapsed_seconds": os_.duration_seconds()})
    return jsonify({"active": False})

@app.route("/api/punch/in", methods=["POST"])
@login_required
def punch_in():
    eid = session.get("emp_id")
    if not eid: return jsonify({"error": "No employee in session"}), 400
    b = request.json or {}; store = get_store(); d = store.load()
    if d.get_open_session_for_employee(eid):
        return jsonify({"error": "Already punched in. Punch out first."}), 400
    pid = (b.get("project_id") or "").strip()
    if not pid: return jsonify({"error": "Project is required"}), 400
    mod = normalize_module(b.get("module") or "")
    proj = d.project_by_id(pid)
    if proj and not proj.use_module: mod = ""
    s = WorkSession(id=short_id(), employee_id=eid, project_id=pid, module=mod,
                    task_category=(b.get("task_category") or "Development").strip(),
                    remark=(b.get("remark") or "").strip(), punch_in=now())
    store.upsert_session(s)
    return jsonify({"ok": True, "session_id": s.id, "punch_in": fmt_dt(s.punch_in)})

@app.route("/api/punch/out", methods=["POST"])
@login_required
def punch_out():
    eid = session.get("emp_id")
    if not eid: return jsonify({"error": "No employee in session"}), 400
    store = get_store(); d = store.load()
    os_ = d.get_open_session_for_employee(eid)
    if not os_: return jsonify({"error": "No active session"}), 400
    os_.punch_out = now(); store.upsert_session(os_)
    return jsonify({"ok": True, "duration": seconds_to_hhmmss(os_.duration_seconds()),
                    "punch_out": fmt_dt(os_.punch_out)})

@app.route("/api/sessions/my")
@login_required
def my_sessions():
    eid = session.get("emp_id")
    if not eid: return jsonify([])
    d_from = parse_date(request.args.get("from") or "") or (date.today() - timedelta(days=7))
    d_to   = parse_date(request.args.get("to")   or "") or date.today()
    d = get_store().load(); pm = {p.id: p for p in d.projects}; out = []
    for s in d.sessions:
        if s.employee_id != eid: continue
        if s.punch_in.date() < d_from or s.punch_in.date() > d_to: continue
        proj = pm.get(s.project_id)
        out.append({"id": s.id, "project": proj.name if proj else s.project_id,
                    "project_code": proj.code if proj else "", "module": s.module,
                    "task_category": s.task_category, "remark": s.remark,
                    "punch_in": fmt_dt(s.punch_in),
                    "punch_out": fmt_dt(s.punch_out) if s.punch_out else None,
                    "is_open": s.is_open,
                    "duration": seconds_to_hhmmss(int((s.punch_out - s.punch_in).total_seconds())) if s.punch_out else "active"})
    return jsonify(sorted(out, key=lambda x: x["punch_in"], reverse=True))

@app.route("/api/password/change", methods=["POST"])
@login_required
def change_password():
    if session.get("role") == "admin":
        return jsonify({"error": "Admin password is fixed"}), 400
    b = request.json or {}; lid = session.get("user"); store = get_store()
    if not store.verify_password(lid, b.get("old_password") or ""):
        return jsonify({"error": "Current password incorrect"}), 400
    np = b.get("new_password") or ""
    if len(np) < 4: return jsonify({"error": "Password must be at least 4 characters"}), 400
    store.set_password(lid, np)
    return jsonify({"ok": True})

# ── Admin: Employees ───────────────────────────────────────────────────────────

@app.route("/api/admin/employees")
@login_required
@admin_required
def list_employees():
    d = get_store().load()
    return jsonify([{"id": e.id, "name": e.name, "emp_code": e.emp_code,
                     "laptop_brand": e.laptop_brand, "laptop_no": e.laptop_no}
                    for e in d.employees])

@app.route("/api/admin/employees", methods=["POST"])
@login_required
@admin_required
def add_employee():
    b = request.json or {}
    try:
        eid = get_store().add_employee(name=b.get("name") or "", emp_code=b.get("emp_code") or "",
                                       laptop_brand=b.get("laptop_brand") or "",
                                       laptop_no=b.get("laptop_no") or "",
                                       email=b.get("email") or "")
        return jsonify({"ok": True, "id": eid})
    except ValueError as e: return jsonify({"error": str(e)}), 400

@app.route("/api/admin/employees/<eid>", methods=["PUT"])
@login_required
@admin_required
def update_employee(eid):
    b = request.json or {}
    try:
        get_store().update_employee(eid, b.get("name") or "", b.get("emp_code") or "", b.get("email") or "")
        return jsonify({"ok": True})
    except ValueError as e: return jsonify({"error": str(e)}), 400

@app.route("/api/admin/employees/<eid>", methods=["DELETE"])
@login_required
@admin_required
def delete_employee(eid):
    try: get_store().delete_employee(eid); return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/admin/employees/<eid>/reset-password", methods=["POST"])
@login_required
@admin_required
def reset_password(eid):
    b = request.json or {}; store = get_store(); d = store.load()
    emp = next((e for e in d.employees if e.id == eid), None)
    if not emp: return jsonify({"error": "Employee not found"}), 404
    store.set_password(emp.emp_code, b.get("password") or "Mach@123")
    return jsonify({"ok": True})

# ── Admin: Projects ────────────────────────────────────────────────────────────

@app.route("/api/admin/projects")
@login_required
@admin_required
def list_projects():
    d = get_store().load()
    return jsonify([{"id": p.id, "code": p.code, "name": p.name, "use_module": p.use_module,
                     "allowed_tasks": p.allowed_tasks, "planned_hours": p.planned_hours}
                    for p in d.projects])

@app.route("/api/admin/projects", methods=["POST"])
@login_required
@admin_required
def add_project():
    b = request.json or {}
    try:
        pid = get_store().add_project(code=b.get("code") or "", name=b.get("name") or "",
                                      use_module=int(b.get("use_module", 1)),
                                      allowed_tasks=b.get("allowed_tasks") or "",
                                      planned_hours=float(b.get("planned_hours") or 0))
        return jsonify({"ok": True, "id": pid})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/admin/projects/<pid>", methods=["PUT"])
@login_required
@admin_required
def update_project(pid):
    b = request.json or {}
    try:
        store = get_store()
        code  = (b.get("code") or "").strip().upper()
        name  = (b.get("name") or "").strip()
        if not code or not name: return jsonify({"error": "Code and name are required"}), 400
        cur = store.conn.cursor()
        cur.execute("UPDATE projects SET code=?, name=? WHERE id=?", (code, name, pid))
        store.conn.commit()
        store.update_project_plans(project_id=pid, planned_hours=float(b.get("planned_hours") or 0),
                                   use_module=bool(int(b.get("use_module", 1))),
                                   allowed_tasks=b.get("allowed_tasks") or "")
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/admin/projects/<pid>", methods=["DELETE"])
@login_required
@admin_required
def delete_project(pid):
    try: get_store().delete_project(pid); return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 400

# ── Admin: Sessions ────────────────────────────────────────────────────────────

@app.route("/api/admin/sessions")
@login_required
@admin_required
def admin_sessions():
    d_from = parse_date(request.args.get("from") or "") or (date.today() - timedelta(days=7))
    d_to   = parse_date(request.args.get("to")   or "") or date.today()
    ef = request.args.get("emp_id") or ""
    d = get_store().load(); em = {e.id: e for e in d.employees}; pm = {p.id: p for p in d.projects}
    out = []
    for s in d.sessions:
        if ef and s.employee_id != ef: continue
        if s.punch_in.date() < d_from or s.punch_in.date() > d_to: continue
        e = em.get(s.employee_id); p = pm.get(s.project_id)
        out.append({"id": s.id, "employee": e.name if e else s.employee_id,
                    "employee_id": s.employee_id, "project": p.name if p else s.project_id,
                    "project_id": s.project_id, "module": s.module,
                    "task_category": s.task_category, "remark": s.remark,
                    "punch_in": fmt_dt(s.punch_in),
                    "punch_out": fmt_dt(s.punch_out) if s.punch_out else None,
                    "is_open": s.is_open,
                    "duration": seconds_to_hhmmss(int((s.punch_out - s.punch_in).total_seconds())) if s.punch_out else "active"})
    return jsonify(sorted(out, key=lambda x: (x["employee"], x["punch_in"])))

@app.route("/api/admin/sessions", methods=["POST"])
@login_required
@admin_required
def admin_add_session():
    b = request.json or {}
    try:
        eid = (b.get("employee_id") or "").strip()
        pid = (b.get("project_id")  or "").strip()
        pi  = parse_user_datetime(b.get("punch_in")  or "")
        po  = parse_user_datetime(b.get("punch_out") or "") if b.get("punch_out") else None
        if not eid or not pid or not pi:
            return jsonify({"error": "employee_id, project_id and punch_in required"}), 400
        s = WorkSession(id=short_id(), employee_id=eid, project_id=pid,
                        module=normalize_module(b.get("module") or ""),
                        task_category=(b.get("task_category") or "Development").strip(),
                        remark=(b.get("remark") or "").strip(), punch_in=pi, punch_out=po)
        get_store().upsert_session(s)
        return jsonify({"ok": True, "id": s.id})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/admin/sessions/<sid>", methods=["PUT"])
@login_required
@admin_required
def admin_update_session(sid):
    b = request.json or {}
    try:
        store = get_store(); d = store.load()
        s = next((x for x in d.sessions if x.id == sid), None)
        if not s: return jsonify({"error": "Session not found"}), 404
        if "employee_id"   in b: s.employee_id   = b["employee_id"]
        if "project_id"    in b: s.project_id    = b["project_id"]
        if "module"        in b: s.module        = normalize_module(b["module"])
        if "task_category" in b: s.task_category = b["task_category"]
        if "remark"        in b: s.remark        = b["remark"]
        if "punch_in"      in b:
            dt = parse_user_datetime(b["punch_in"])
            if dt: s.punch_in = dt
        if "punch_out"     in b:
            s.punch_out = parse_user_datetime(b["punch_out"]) if b["punch_out"] else None
        store.upsert_session(s)
        return jsonify({"ok": True})
    except Exception as e: return jsonify({"error": str(e)}), 400

@app.route("/api/admin/sessions/<sid>", methods=["DELETE"])
@login_required
@admin_required
def admin_delete_session(sid):
    get_store().delete_session(sid)
    return jsonify({"ok": True})

# ── Admin: Summary ─────────────────────────────────────────────────────────────

@app.route("/api/admin/summary")
@login_required
@admin_required
def admin_summary():
    d_from = parse_date(request.args.get("from") or "") or start_of_week(date.today())
    d_to   = parse_date(request.args.get("to")   or "") or date.today()
    store = get_store(); d = store.load()
    em = {e.id: e for e in d.employees}; pm = {p.id: p for p in d.projects}
    pp = {p.name: p.planned_hours for p in d.projects}; pu = {}; epm = {}
    for s in d.sessions:
        if s.is_open or not s.punch_out: continue
        sd = s.punch_in.date()
        if sd < d_from or sd > d_to: continue
        p = pm.get(s.project_id); pn = p.name if p else s.project_id
        e = em.get(s.employee_id); en = e.name if e else s.employee_id
        # Use explicit punch_out - punch_in to avoid duration_seconds() falling back to now()
        raw_sec = max(0, int((s.punch_out - s.punch_in).total_seconds()))
        h = raw_sec / 3600.0
        pu[pn] = pu.get(pn, 0.0) + h
        k = (en, s.employee_id, pn, s.module or "-"); epm[k] = epm.get(k, 0.0) + h
    def _hhmm(h):
        sec = int(h * 3600)
        return f"{sec//3600:02d}:{(sec%3600)//60:02d}"

    ps = [{"project": n,
           "planned_hours": round(pp.get(n, 0.0), 2),
           "planned_hhmm": _hhmm(pp.get(n, 0.0)),
           "used_hours": round(pu.get(n, 0.0), 2),
           "used_hhmm": _hhmm(pu.get(n, 0.0)),
           "remaining": round(pp.get(n, 0.0) - pu.get(n, 0.0), 2) if pp.get(n, 0.0) > 0 else None,
           "remaining_hhmm": _hhmm(max(0, pp.get(n, 0.0) - pu.get(n, 0.0))) if pp.get(n, 0.0) > 0 else None}
          for n in sorted(set(list(pp) + list(pu)), key=str.lower)]
    es = [{"employee": en, "employee_id": eid, "project": pn, "module": mod,
           "hours": round(h, 2), "hours_hhmm": _hhmm(h),
           "approved": store.is_approved(eid, d_from, d_to)}
          for (en, eid, pn, mod), h in sorted(epm.items())]
    return jsonify({"date_from": fmt_date(d_from), "date_to": fmt_date(d_to),
                    "project_summary": ps, "employee_summary": es})

# ── Admin: SOI Plan ────────────────────────────────────────────────────────────

@app.route("/api/admin/projects/<pid>/soi", methods=["GET"])
@login_required
@admin_required
def get_soi_plan(pid):
    rows = get_store().get_project_soi_plan(pid)
    return jsonify([{"task_name": t, "soi_level": s, "planned_pct": p} for t, s, p in rows])

@app.route("/api/admin/projects/<pid>/modules", methods=["GET"])
@login_required
@admin_required
def get_project_modules_route(pid):
    rows = get_store().get_project_modules(pid)
    return jsonify([{"module_name": m, "planned_hours": h} for m, h in rows])

@app.route("/api/admin/projects/<pid>/modules", methods=["POST"])
@login_required
@admin_required
def set_project_modules_route(pid):
    b = request.json or {}
    rows_raw = b.get("modules") or []
    modules = []
    for row in rows_raw:
        name = (row.get("module_name") or "").strip()
        hrs  = float(row.get("planned_hours") or 0.0)
        if name:
            modules.append((name, hrs))
    get_store().set_project_modules(pid, modules)
    return jsonify({"ok": True})

@app.route("/api/admin/projects/<pid>/soi", methods=["POST"])
@login_required
@admin_required
def set_soi_plan(pid):
    b = request.json or {}
    rows_raw = b.get("rows") or []
    rows = []
    for r in rows_raw:
        task = (r.get("task_name") or "").strip()
        soi  = int(r.get("soi_level") or 1)
        pct  = float(r.get("planned_pct") or 0.0)
        if task and soi in (1, 2, 3, 4):
            rows.append((task, soi, pct))
    get_store().set_project_soi_plan(pid, rows)
    return jsonify({"ok": True})

# ── Admin: Approvals ───────────────────────────────────────────────────────────

@app.route("/api/admin/project-totals")
@login_required
@admin_required
def project_totals():
    """All-time hours per project/module/task with SOI. Only completed sessions."""
    store = get_store(); d = store.load(); pm = {p.id: p for p in d.projects}
    proj_sec = {}; mod_sec = {}; modtask_sec = {}
    for s in d.sessions:
        if not s.punch_out: continue
        raw_sec = max(0, int((s.punch_out - s.punch_in).total_seconds()))
        p   = pm.get(s.project_id); pn = p.name if p else s.project_id
        mod = (s.module or "General").strip() or "General"
        tsk = (s.task_category or "—").strip() or "—"
        proj_sec[pn]                = proj_sec.get(pn, 0) + raw_sec
        mod_sec[(pn, mod)]          = mod_sec.get((pn, mod), 0) + raw_sec
        modtask_sec[(pn, mod, tsk)] = modtask_sec.get((pn, mod, tsk), 0) + raw_sec
    def _fmt(sec):
        sec = max(0, int(sec))
        return "%02d:%02d" % (sec // 3600, (sec % 3600) // 60)
    projects = []; seen = set()
    for p in sorted(d.projects, key=lambda x: x.name.lower()):
        seen.add(p.name)
        sec      = proj_sec.get(p.name, 0)
        planned  = p.planned_hours or 0
        plan_sec = int(planned * 3600)
        pct      = min(100, round(sec / plan_sec * 100)) if plan_sec > 0 else None
        rem_sec  = max(0, plan_sec - sec) if plan_sec > 0 else None
        # SOI map: task_name -> (soi_level, planned_pct)
        soi_rows = store.get_project_soi_plan(p.id)
        soi_plan = [{"task_name": t, "soi_level": sl, "planned_pct": pc} for t, sl, pc in soi_rows]
        soi_map  = {t: (sl, pc) for t, sl, pc in soi_rows}
        # Module breakdown
        mods = []
        mod_keys = sorted(set(mod for (pn2, mod) in mod_sec if pn2 == p.name))
        for mod in mod_keys:
            msec = mod_sec.get((p.name, mod), 0)
            tasks = []
            for (pn2, mod2, tsk), tsec in sorted(modtask_sec.items()):
                if pn2 != p.name or mod2 != mod: continue
                sl, pc = soi_map.get(tsk, (None, 0.0))
                t_plan = int(plan_sec * pc / 100.0) if plan_sec > 0 and pc > 0 else 0
                t_rem  = max(0, t_plan - tsec) if t_plan > 0 else None
                tasks.append({"task": tsk, "soi_level": sl,
                    "used_hhmm": _fmt(tsec), "used_sec": tsec,
                    "planned_hhmm": _fmt(t_plan) if t_plan else None, "planned_sec": t_plan,
                    "remaining_hhmm": _fmt(t_rem) if t_rem is not None else None, "remaining_sec": t_rem})
            mods.append({"module": mod, "seconds": msec, "display": _fmt(msec),
                         "used_hhmm": _fmt(msec),
                         "planned_hhmm": None, "remaining_hhmm": None, "tasks": tasks})
        projects.append({"id": p.id, "name": p.name, "code": p.code,
            "planned_hours": planned, "planned_hhmm": _fmt(plan_sec),
            "used_seconds": sec, "used_hhmm": _fmt(sec),
            "used_hours": round(sec/3600, 2), "pct": pct,
            "remaining_hhmm": _fmt(rem_sec) if rem_sec is not None else None,
            "remaining_sec": rem_sec,
            "modules": mods, "soi_plan": soi_plan})
    for pn, sec in sorted(proj_sec.items()):
        if pn in seen: continue
        mods = [{"module": mod, "seconds": msec, "display": _fmt(msec),
                 "used_hhmm": _fmt(msec), "tasks": []}
                for (pn2, mod), msec in sorted(mod_sec.items()) if pn2 == pn]
        projects.append({"id": None, "name": pn, "code": "?",
            "planned_hours": 0, "planned_hhmm": "00:00", "remaining_hhmm": None,
            "used_seconds": sec, "used_hhmm": _fmt(sec),
            "used_hours": round(sec/3600, 2), "pct": None,
            "modules": mods, "soi_plan": []})
    return jsonify(projects)

@app.route("/api/admin/approvals")
@login_required
@admin_required
def get_approvals():
    d_from = parse_date(request.args.get("from") or "") or start_of_week(date.today())
    d_to   = parse_date(request.args.get("to")   or "") or date.today()
    store = get_store(); d = store.load()
    return jsonify([{"employee_id": e.id, "name": e.name,
                     "approved": store.is_approved(e.id, d_from, d_to)} for e in d.employees])

@app.route("/api/admin/approvals", methods=["POST"])
@login_required
@admin_required
def set_approval():
    b = request.json or {}
    eid = (b.get("employee_id") or "").strip()
    df  = parse_date(b.get("date_from") or "")
    dt  = parse_date(b.get("date_to")   or "")
    if not eid or not df or not dt:
        return jsonify({"error": "employee_id, date_from and date_to are required"}), 400
    get_store().set_approval(eid, df, dt, bool(b.get("approved", False)))
    return jsonify({"ok": True})

# ── Admin: Export ──────────────────────────────────────────────────────────────


def _export_all_xlsx(store, out_path: str, d_from, d_to, approved_ids: set) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    import datetime as _dt2

    d = store.load()
    proj_name    = {p.id: p.name or p.code or p.id for p in d.projects}
    proj_planned = {p.id: float(getattr(p, 'planned_hours', 0) or 0) for p in d.projects}
    emp_name     = {e.id: e.name or e.id for e in d.employees}
    emp_code     = {e.id: getattr(e, 'emp_code', '') or '' for e in d.employees}

    # ── Colour Palette ────────────────────────────────────────────────────────
    # Deep navy header
    C_NAVY        = "0D2137"
    C_NAVY_MID    = "163354"
    C_NAVY_LIGHT  = "1E4976"
    # Accent teal/blue
    C_ACCENT      = "0E7490"
    C_ACCENT_SOFT = "CFFAFE"
    # Greens
    C_GREEN       = "065F46"
    C_GREEN_SOFT  = "D1FAE5"
    C_GREEN_MID   = "6EE7B7"
    # Ambers
    C_AMBER       = "92400E"
    C_AMBER_SOFT  = "FEF3C7"
    # Reds
    C_RED         = "991B1B"
    C_RED_SOFT    = "FEE2E2"
    # Rows
    C_ROW_A       = "F0F9FF"   # lightest sky
    C_ROW_B       = "FFFFFF"
    C_SUBTTL      = "E0F2FE"   # subtotal row – light cyan
    C_GRAND       = "BAE6FD"   # grand total  – medium cyan
    C_TITLE_BG    = "0D2137"
    C_TITLE_BAND  = "164E63"   # secondary banner strip
    C_WHITE       = "FFFFFF"

    # Thin/medium borders
    thin   = Side(style="thin",   color="CBD5E1")
    medium = Side(style="medium", color="0D2137")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    MBDR   = Border(left=medium, right=medium, top=medium, bottom=medium)
    BOT_M  = Border(left=thin, right=thin, top=thin, bottom=medium)

    # Alignments
    CC = Alignment(horizontal="center", vertical="center")
    LC = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    RC = Alignment(horizontal="right",  vertical="center")

    def fill(hex_):   return PatternFill("solid", fgColor=hex_)
    def font(hex_, bold=False, sz=10, name="Calibri", italic=False):
        return Font(color=hex_, bold=bold, size=sz, name=name, italic=italic)

    def sec_to_hhmm(sec):
        sec = max(0, int(sec))
        return "%02d:%02d" % (sec // 3600, (sec % 3600) // 60)

    def cell_set(ws, row, col, val, bg, fg, bold=False, align=CC, sz=10,
                 border=BORDER, italic=False, number_format=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill      = fill(bg)
        cell.font      = Font(color=fg, bold=bold, size=sz, name="Calibri", italic=italic)
        cell.alignment = align
        cell.border    = border
        if number_format: cell.number_format = number_format
        return cell

    def header_row(ws, row, cols):
        ws.row_dimensions[row].height = 26
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill      = fill(C_NAVY)
            cell.font      = Font(color=C_WHITE, bold=True, size=11, name="Calibri")
            cell.alignment = CC
            cell.border    = MBDR

    def data_row(ws, row, values, cols_cfg):
        # cols_cfg: list of (bg_hex, fg_hex, bold, align)
        ws.row_dimensions[row].height = 18
        alt = (row % 2 == 0)
        for c, (val, cfg) in enumerate(zip(values, cols_cfg), 1):
            bg, fg, bold, align = cfg
            if alt and bg in (C_ROW_B,):
                bg = C_ROW_A
            cell_set(ws, row, c, val, bg, fg, bold, align)

    def subtotal_row(ws, row, values, ncols):
        ws.row_dimensions[row].height = 20
        for c, val in enumerate(values, 1):
            cell_set(ws, row, c, val, C_SUBTTL, C_ACCENT, True, CC if c > 1 else LC, 10, BOT_M)
        for c in range(len(values)+1, ncols+1):
            cell_set(ws, row, c, "", C_SUBTTL, C_ACCENT, False, CC, 10, BOT_M)

    def grand_total_row(ws, row, values, ncols):
        ws.row_dimensions[row].height = 22
        for c, val in enumerate(values, 1):
            cell_set(ws, row, c, val, C_GRAND, C_NAVY, True, CC if c > 1 else LC, 11, MBDR)
        for c in range(len(values)+1, ncols+1):
            cell_set(ws, row, c, "", C_GRAND, C_NAVY, False, CC, 11, MBDR)

    def banner(ws, sheet_title, subtitle, ncols):
        # Row 1 – dark title bar
        ws.row_dimensions[1].height = 38
        cell = ws.cell(row=1, column=1,
                       value=f"  PULSE TRACKER  ·  {sheet_title.upper()}")
        cell.fill = fill(C_TITLE_BG)
        cell.font = Font(color=C_WHITE, bold=True, size=15, name="Calibri")
        cell.alignment = LC
        cell.border = MBDR
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        # Row 2 – teal info bar
        ws.row_dimensions[2].height = 20
        cell2 = ws.cell(row=2, column=1, value=f"  {subtitle}")
        cell2.fill = fill(C_TITLE_BAND)
        cell2.font = Font(color="BAE6FD", size=9, name="Calibri", italic=True)
        cell2.alignment = LC
        cell2.border = Border(left=medium, right=medium, bottom=thin)
        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        # Row 3 – spacer
        ws.row_dimensions[3].height = 5
        for c in range(1, ncols+1):
            ws.cell(row=3, column=c).fill = fill("E0F2FE")

    def closed_sessions():
        for s in d.sessions:
            if not getattr(s, 'punch_out', None): continue
            sd2 = s.punch_in.date()
            if sd2 < d_from or sd2 > d_to: continue
            yield s

    wb = Workbook()
    try: wb.remove(wb.active)
    except: pass

    period   = f"{d_from.strftime('%d %b %Y')}  →  {d_to.strftime('%d %b %Y')}"
    gen_date = f"Generated: {_dt2.date.today().strftime('%d %b %Y')}"
    subtitle_line = f"Period: {period}   |   {gen_date}"

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — PROJECT SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Project Summary")
    banner(ws1, "Project Summary", subtitle_line, 5)
    header_row(ws1, 4, ["Project", "Planned (HH:MM)", "Used (HH:MM)", "Remaining (HH:MM)", "% Used"])
    for i, w in enumerate([36, 18, 18, 18, 12], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A5"

    proj_used_sec = {}
    for s in closed_sessions():
        proj_used_sec[s.project_id] = proj_used_sec.get(s.project_id, 0) +             max(0, int((s.punch_out - s.punch_in).total_seconds()))

    all_pids = sorted(set(list(proj_planned) + list(proj_used_sec)),
                      key=lambda pid: proj_name.get(pid, '').lower())
    r = 5; total_used_s = 0; total_plan_s = 0
    for pid in all_pids:
        pn       = proj_name.get(pid, pid)
        plan_sec = int(proj_planned.get(pid, 0) * 3600)
        used_sec = proj_used_sec.get(pid, 0)
        if used_sec == 0 and plan_sec == 0: continue
        rem_sec  = max(0, plan_sec - used_sec) if plan_sec > 0 else 0
        pct      = round(min(100, used_sec / plan_sec * 100), 1) if plan_sec > 0 else None
        total_used_s += used_sec; total_plan_s += plan_sec
        alt = (r % 2 == 0)
        row_bg = C_ROW_A if alt else C_ROW_B

        cell_set(ws1, r, 1, pn, row_bg, C_NAVY, True, LC)
        cell_set(ws1, r, 2, sec_to_hhmm(plan_sec) if plan_sec else "—", row_bg, "374151", False, CC)
        cell_set(ws1, r, 3, sec_to_hhmm(used_sec), row_bg, "0E4C8E", True, CC)

        # Remaining – colour-coded
        rem_val = sec_to_hhmm(rem_sec) if plan_sec else "—"
        if plan_sec and rem_sec == 0:
            cell_set(ws1, r, 4, rem_val, C_RED_SOFT, C_RED, True, CC)
        elif plan_sec:
            cell_set(ws1, r, 4, rem_val, C_GREEN_SOFT, C_GREEN, True, CC)
        else:
            cell_set(ws1, r, 4, rem_val, row_bg, "94A3B8", False, CC)

        # % Used – traffic light
        pct_str = f"{pct}%" if pct is not None else "—"
        if pct is None:
            cell_set(ws1, r, 5, pct_str, row_bg, "94A3B8", False, CC)
        elif pct >= 100:
            cell_set(ws1, r, 5, pct_str, C_RED_SOFT, C_RED, True, CC)
        elif pct >= 80:
            cell_set(ws1, r, 5, pct_str, C_AMBER_SOFT, C_AMBER, True, CC)
        else:
            cell_set(ws1, r, 5, pct_str, C_GREEN_SOFT, C_GREEN, True, CC)
        r += 1

    if r > 5:
        tot_pct = round(min(100, total_used_s / total_plan_s * 100), 1) if total_plan_s > 0 else None
        grand_total_row(ws1, r, [
            "GRAND TOTAL",
            sec_to_hhmm(total_plan_s) if total_plan_s else "—",
            sec_to_hhmm(total_used_s),
            sec_to_hhmm(max(0, total_plan_s - total_used_s)) if total_plan_s else "—",
            f"{tot_pct}%" if tot_pct is not None else "—"
        ], 5)
    else:
        ws1.merge_cells("A5:E5")
        cell_set(ws1, 5, 1, "No sessions recorded in this date range", C_ROW_A, "94A3B8", False, LC)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — PROJECT + MODULE
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Project + Module")
    banner(ws2, "Project + Module Breakdown", subtitle_line, 4)
    header_row(ws2, 4, ["Project", "Module", "Task Category", "Used (HH:MM)"])
    for i, w in enumerate([36, 20, 24, 18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A5"

    pm_used = {}
    for s in closed_sessions():
        pn   = proj_name.get(s.project_id, s.project_id)
        mod  = getattr(s, 'module', '') or 'General'
        task = getattr(s, 'task_category', '') or '—'
        dur  = max(0, int((s.punch_out - s.punch_in).total_seconds()))
        pm_used[(pn, mod, task)] = pm_used.get((pn, mod, task), 0) + dur

    # Assign a distinct accent colour per project
    proj_palette = [
        ("E0F2FE", "0C4A6E"),  # sky
        ("F0FDF4", "14532D"),  # mint
        ("FDF4FF", "581C87"),  # purple
        ("FFF7ED", "7C2D12"),  # orange
        ("F0F9FF", "0C4A6E"),  # repeat
    ]
    proj_colour_map = {}
    colour_idx = 0

    r = 5; last_proj = None; proj_sub_s = 0; total2_s = 0
    for key in sorted(pm_used, key=lambda k: (k[0].lower(), k[1].lower(), k[2].lower())):
        pn, mod, task = key
        sec = pm_used[key]; total2_s += sec

        if pn != last_proj:
            if last_proj is not None:
                subtotal_row(ws2, r, [f"  {last_proj}  —  Sub Total", "", "", sec_to_hhmm(proj_sub_s)], 4)
                r += 1
            if pn not in proj_colour_map:
                proj_colour_map[pn] = proj_palette[colour_idx % len(proj_palette)]
                colour_idx += 1
            last_proj = pn; proj_sub_s = 0

        proj_sub_s += sec
        pb, pf = proj_colour_map[pn]
        alt = (r % 2 == 0)
        row_bg = pb if not alt else ("EFF6FF" if pb == "E0F2FE" else pb)

        cell_set(ws2, r, 1, pn,   row_bg, pf, True, LC)
        cell_set(ws2, r, 2, mod,  row_bg, "374151", False, LC)
        cell_set(ws2, r, 3, task, row_bg, "374151", False, LC)
        cell_set(ws2, r, 4, sec_to_hhmm(sec), row_bg, "0E4C8E", True, CC)
        r += 1

    if last_proj:
        subtotal_row(ws2, r, [f"  {last_proj}  —  Sub Total", "", "", sec_to_hhmm(proj_sub_s)], 4)
        r += 1

    if r > 5:
        grand_total_row(ws2, r, ["GRAND TOTAL", "", "", sec_to_hhmm(total2_s)], 4)
    else:
        ws2.merge_cells("A5:D5")
        cell_set(ws2, 5, 1, "No sessions recorded in this date range", C_ROW_A, "94A3B8", False, LC)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — EMPLOYEE SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Employee Summary")
    banner(ws3, "Employee Summary", subtitle_line, 7)
    header_row(ws3, 4, ["Employee", "Emp Code", "Project", "Module", "Task", "Used (HH:MM)", "Approved"])
    for i, w in enumerate([26, 14, 32, 18, 22, 18, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A5"

    emp_used = {}
    for s in closed_sessions():
        en   = emp_name.get(s.employee_id, s.employee_id)
        ec   = emp_code.get(s.employee_id, '')
        pn   = proj_name.get(s.project_id, s.project_id)
        mod  = getattr(s, 'module', '') or 'General'
        task = getattr(s, 'task_category', '') or '—'
        appr = s.employee_id in approved_ids
        dur  = max(0, int((s.punch_out - s.punch_in).total_seconds()))
        emp_used[(en, ec, pn, mod, task, appr)] = emp_used.get((en, ec, pn, mod, task, appr), 0) + dur

    # Distinct colour per employee
    emp_palette = [
        ("EFF6FF", "1E3A8A"),
        ("FDF4FF", "581C87"),
        ("FFF7ED", "7C2D12"),
        ("F0FDF4", "14532D"),
        ("FFF1F2", "881337"),
    ]
    emp_colour_map = {}
    emp_colour_idx = 0

    r = 5; last_emp = None; emp_sub_s = 0; total3_s = 0
    for key in sorted(emp_used, key=lambda k: (k[0].lower(), k[2].lower())):
        en, ec, pn, mod, task, appr = key
        sec = emp_used[key]; total3_s += sec

        if en != last_emp:
            if last_emp:
                subtotal_row(ws3, r, [f"  {last_emp}  —  Sub Total", "", "", "", "", sec_to_hhmm(emp_sub_s), ""], 7)
                r += 1
            if en not in emp_colour_map:
                emp_colour_map[en] = emp_palette[emp_colour_idx % len(emp_palette)]
                emp_colour_idx += 1
            last_emp = en; emp_sub_s = 0

        emp_sub_s += sec
        eb, ef = emp_colour_map[en]

        cell_set(ws3, r, 1, en,   eb, ef, True, LC)
        cell_set(ws3, r, 2, ec,   eb, "374151", False, CC)
        cell_set(ws3, r, 3, pn,   eb, "374151", False, LC)
        cell_set(ws3, r, 4, mod,  eb, "374151", False, LC)
        cell_set(ws3, r, 5, task, eb, "374151", False, LC)
        cell_set(ws3, r, 6, sec_to_hhmm(sec), eb, "0E4C8E", True, CC)
        if appr:
            cell_set(ws3, r, 7, "✓  Approved", C_GREEN_SOFT, C_GREEN, True, CC)
        else:
            cell_set(ws3, r, 7, "Pending", C_AMBER_SOFT, C_AMBER, False, CC)
        r += 1

    if last_emp:
        subtotal_row(ws3, r, [f"  {last_emp}  —  Sub Total", "", "", "", "", sec_to_hhmm(emp_sub_s), ""], 7)
        r += 1
    if r > 5:
        grand_total_row(ws3, r, ["GRAND TOTAL", "", "", "", "", sec_to_hhmm(total3_s), ""], 7)
    else:
        ws3.merge_cells("A5:G5")
        cell_set(ws3, 5, 1, "No sessions recorded in this date range", C_ROW_A, "94A3B8", False, LC)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — SOI SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("SOI Summary")
    banner(ws4, "SOI Summary", subtitle_line, 7)
    header_row(ws4, 4, ["Project", "SOI Level", "Task", "Planned (HH:MM)", "Used (HH:MM)", "Remaining (HH:MM)", "% Used"])
    for i, w in enumerate([32, 12, 24, 18, 18, 18, 12], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A5"

    # SOI level accent colours
    soi_colours = {
        "SOI-1": ("EFF6FF", "1D4ED8"),
        "SOI-2": ("F0FDF4", "15803D"),
        "SOI-3": ("FFF7ED", "C2410C"),
        "SOI-4": ("FFF1F2", "BE123C"),
        "—":     ("F8FAFC", "64748B"),
    }

    soi_map = {}
    for p in d.projects:
        try:
            for task, soi_level, planned_pct in store.get_project_soi_plan(p.id):
                if task:
                    soi_map[(p.id, task)] = (int(soi_level or 1), float(planned_pct or 0))
        except: pass

    soi_used = {}
    for s in closed_sessions():
        task  = getattr(s, 'task_category', '') or '—'
        sl, _ = soi_map.get((s.project_id, task), (None, 0))
        pn    = proj_name.get(s.project_id, s.project_id)
        label = f"SOI-{sl}" if sl else "—"
        dur   = max(0, int((s.punch_out - s.punch_in).total_seconds()))
        soi_used[(pn, s.project_id, label, task)] = soi_used.get((pn, s.project_id, label, task), 0) + dur

    sorted_soi = sorted(soi_used.keys(), key=lambda k: (k[0].lower(), k[2], k[3].lower()))

    r = 5; total4_s = 0
    idx = 0
    while idx < len(sorted_soi):
        pn, pid, label, task = sorted_soi[idx]
        # Find group with same project + SOI level
        j = idx
        while j < len(sorted_soi) and sorted_soi[j][0] == pn and sorted_soi[j][2] == label:
            j += 1
        group = sorted_soi[idx:j]
        group_start = r
        sb, sf = soi_colours.get(label, ("F8FAFC", "374151"))

        for key2 in group:
            pn2, pid2, label2, task2 = key2
            used_sec = soi_used[key2]
            _, pct_plan = soi_map.get((pid2, task2), (None, 0))
            plan_sec = int(proj_planned.get(pid2, 0) * 3600 * pct_plan / 100) if pct_plan else 0
            rem_sec  = max(0, plan_sec - used_sec) if plan_sec else 0
            pct_u    = round(min(100, used_sec / plan_sec * 100), 1) if plan_sec else None
            total4_s += used_sec
            alt = (r % 2 == 0)
            row_bg = sb if not alt else ("E0EEFF" if sb == "EFF6FF" else sb)

            cell_set(ws4, r, 1, pn2,  row_bg, C_NAVY, True, LC)
            # SOI Level cell placeholder – merged after loop
            cell_set(ws4, r, 2, label2 if r == group_start else "", sb, sf, True, CC)
            cell_set(ws4, r, 3, task2, row_bg, "374151", False, LC)
            cell_set(ws4, r, 4, sec_to_hhmm(plan_sec) if plan_sec else "—", row_bg, "374151", False, CC)
            cell_set(ws4, r, 5, sec_to_hhmm(used_sec), row_bg, "0E4C8E", True, CC)

            if plan_sec:
                if rem_sec == 0:
                    cell_set(ws4, r, 6, sec_to_hhmm(rem_sec), C_RED_SOFT, C_RED, True, CC)
                else:
                    cell_set(ws4, r, 6, sec_to_hhmm(rem_sec), C_GREEN_SOFT, C_GREEN, True, CC)
            else:
                cell_set(ws4, r, 6, "—", row_bg, "94A3B8", False, CC)

            if pct_u is None:
                cell_set(ws4, r, 7, "—", row_bg, "94A3B8", False, CC)
            elif pct_u >= 100:
                cell_set(ws4, r, 7, f"{pct_u}%", C_RED_SOFT, C_RED, True, CC)
            elif pct_u >= 80:
                cell_set(ws4, r, 7, f"{pct_u}%", C_AMBER_SOFT, C_AMBER, True, CC)
            else:
                cell_set(ws4, r, 7, f"{pct_u}%", C_GREEN_SOFT, C_GREEN, True, CC)
            r += 1

        # Merge SOI Level column for this group
        if len(group) > 1:
            col_ltr = get_column_letter(2)
            ws4.merge_cells(f"{col_ltr}{group_start}:{col_ltr}{r-1}")
            mc = ws4[f"{col_ltr}{group_start}"]
            mc.value     = label
            mc.fill      = fill(sb)
            mc.font      = Font(color=sf, bold=True, size=11, name="Calibri")
            mc.alignment = Alignment(horizontal="center", vertical="center")
            mc.border    = MBDR

        idx = j

    if r > 5:
        grand_total_row(ws4, r, ["GRAND TOTAL", "", "", "", sec_to_hhmm(total4_s), "", ""], 7)
    else:
        ws4.merge_cells("A5:G5")
        cell_set(ws4, 5, 1, "No SOI data in this date range", C_ROW_A, "94A3B8", False, LC)

    # ── Tab colours ───────────────────────────────────────────────────────────
    ws1.sheet_properties.tabColor = "0D2137"
    ws2.sheet_properties.tabColor = "0E7490"
    ws3.sheet_properties.tabColor = "065F46"
    ws4.sheet_properties.tabColor = "6D28D9"

    wb.save(out_path)


# ── Daily Email Scheduler ─────────────────────────────────────────────────────
import smtplib, threading, time as _time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime as _dt, timedelta as _td, date as _date

REPORT_TO   = "krishnamoorthyd@machglobaltech.com"
SMTP_HOST   = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT   = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER   = os.environ.get("SMTP_USER", "")
SMTP_PASS   = os.environ.get("SMTP_PASS", "")
SMTP_FROM   = os.environ.get("SMTP_FROM", SMTP_USER)

def _hhmm_f(h: float) -> str:
    sec = max(0, int(h * 3600))
    return "%02d:%02d" % (sec // 3600, (sec % 3600) // 60)

def _build_email_html(target_date: "_date") -> str:
    store2 = DataStore(DB_PATH)
    try:
        d2 = store2.load()
        em = {e.id: e for e in d2.employees}
        pm = {p.id: p for p in d2.projects}

        # Aggregate per project and per employee
        proj_used_sec: dict = {}   # project_id -> seconds
        emp_total_sec: dict = {}   # emp_name -> seconds  (total for the day)
        emp_proj_sec:  dict = {}   # (emp_name, proj_name, module) -> seconds

        for s in d2.sessions:
            if not s.punch_out: continue
            if s.punch_in.date() != target_date: continue
            pid = s.project_id
            pn  = (pm.get(pid) or type("X",(),{"name":pid})()).name
            en  = (em.get(s.employee_id) or type("X",(),{"name":s.employee_id})()).name
            raw = max(0, int((s.punch_out - s.punch_in).total_seconds()))
            proj_used_sec[pid] = proj_used_sec.get(pid, 0) + raw
            emp_total_sec[en]  = emp_total_sec.get(en, 0) + raw
            k = (en, pn, s.module or "—")
            emp_proj_sec[k] = emp_proj_sec.get(k, 0) + raw

        date_str = target_date.strftime("%d %b %Y")
        acc = "#2563eb"; grn = "#16a34a"; red = "#dc2626"; amb = "#d97706"; bdr = "#e2e8f0"

        def hhmm(sec):
            sec = max(0, int(sec))
            return "%02d:%02d" % (sec // 3600, (sec % 3600) // 60)

        th = f'style="background:{acc};color:#fff;padding:10px 14px;text-align:left;font-size:12px;font-weight:700;border:none"' 
        td = 'style="padding:10px 14px;border-bottom:1px solid #f1f5f9;font-size:13px;color:#374151;vertical-align:middle"'

        total_sec = sum(proj_used_sec.values())

        # ── Project rows: project | planned | used | remaining | % ─────────────
        p_rows = ""
        for p in sorted(d2.projects, key=lambda x: proj_used_sec.get(x.id, 0), reverse=True):
            used_s = proj_used_sec.get(p.id, 0)
            if used_s == 0: continue
            plan_s = int((p.planned_hours or 0) * 3600)
            rem_s  = max(0, plan_s - used_s) if plan_s else None
            pct    = round(min(100, used_s / plan_s * 100)) if plan_s else None
            if pct is not None:
                bar_col = red if pct >= 100 else (amb if pct >= 80 else grn)
                bar = (f'<div style="background:#e2e8f0;border-radius:99px;height:7px;width:110px;display:inline-block;vertical-align:middle">' +
                       f'<div style="background:{bar_col};border-radius:99px;height:7px;width:{min(100,pct)}%;"></div></div>' +
                       f' <span style="font-size:11px;font-weight:700;color:{bar_col}">{pct}%</span>')
            else:
                bar = '<span style="color:#cbd5e1;font-size:12px">No budget</span>'
            rem_col = red if (rem_s == 0 and plan_s) else grn
            p_rows += (
                f'<tr>' +
                f'<td {td}><strong style="color:#0f172a">{p.name}</strong>' +
                (f'<br><span style="font-size:11px;background:#e0e7ff;color:#3730a3;border-radius:4px;padding:1px 6px">{p.code}</span>' if p.code else '') +
                '</td>' +
                f'<td {td} style="font-family:monospace;text-align:center">{hhmm(plan_s) if plan_s else "—"}</td>' +
                f'<td {td} style="font-family:monospace;color:{acc};font-weight:700;text-align:center">{hhmm(used_s)}</td>' +
                f'<td {td} style="font-family:monospace;color:{rem_col};font-weight:700;text-align:center">{hhmm(rem_s) if rem_s is not None else "—"}</td>' +
                f'<td {td}>{bar}</td>' +
                '</tr>'
            )

        # ── Employee total rows: name | total hours ───────────────────────────
        e_total_rows = ""
        for en in sorted(emp_total_sec, key=lambda x: emp_total_sec[x], reverse=True):
            e_total_rows += (
                f'<tr>' +
                f'<td {td}><strong style="color:#0f172a">{en}</strong></td>' +
                f'<td {td} style="font-family:monospace;color:{grn};font-weight:700;text-align:center">{hhmm(emp_total_sec[en])}</td>' +
                '</tr>'
            )

        # ── Employee detail rows: name | project | module | hours ─────────────
        e_detail_rows = ""
        for (en, pn, mod), sec in sorted(emp_proj_sec.items(), key=lambda x: (x[0][0], -x[1])):
            e_detail_rows += (
                f'<tr>' +
                f'<td {td}><strong>{en}</strong></td>' +
                f'<td {td}>{pn}</td>' +
                f'<td {td}>{mod}</td>' +
                f'<td {td} style="font-family:monospace;color:{grn};font-weight:700;text-align:center">{hhmm(sec)}</td>' +
                '</tr>'
            )

        if not p_rows:
            body = f'<p style="color:#94a3b8;text-align:center;padding:32px;font-size:14px">No sessions recorded on {date_str}.</p>'
        else:
            body = f"""
            <div style="display:flex;gap:14px;margin-bottom:28px;flex-wrap:wrap">
              <div style="flex:1;background:#eff6ff;border-radius:12px;padding:16px;text-align:center;min-width:110px">
                <div style="font-size:24px;font-weight:900;color:{acc};font-family:monospace">{hhmm(total_sec)}</div>
                <div style="font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:4px">Total Hours</div>
              </div>
              <div style="flex:1;background:#f0fdf4;border-radius:12px;padding:16px;text-align:center;min-width:110px">
                <div style="font-size:24px;font-weight:900;color:{grn};font-family:monospace">{len(proj_used_sec)}</div>
                <div style="font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:4px">Projects Active</div>
              </div>
              <div style="flex:1;background:#fefce8;border-radius:12px;padding:16px;text-align:center;min-width:110px">
                <div style="font-size:24px;font-weight:900;color:#d97706;font-family:monospace">{len(emp_total_sec)}</div>
                <div style="font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:4px">Employees</div>
              </div>
            </div>

            <h2 style="font-size:14px;font-weight:800;color:#0f172a;margin:0 0 10px;letter-spacing:-.2px">📊 Project Summary</h2>
            <table width="100%" cellpadding="0" cellspacing="0"
              style="border-collapse:collapse;border:1px solid {bdr};border-radius:10px;overflow:hidden;margin-bottom:28px">
              <thead><tr>
                <th {th}>Project</th>
                <th {th} style="text-align:center">Planned</th>
                <th {th} style="text-align:center">Used Today</th>
                <th {th} style="text-align:center">Remaining</th>
                <th {th}>Budget</th>
              </tr></thead>
              <tbody>{p_rows}</tbody>
            </table>

            <h2 style="font-size:14px;font-weight:800;color:#0f172a;margin:0 0 10px">👥 Employee Hours — Today</h2>
            <table width="100%" cellpadding="0" cellspacing="0"
              style="border-collapse:collapse;border:1px solid {bdr};border-radius:10px;overflow:hidden;margin-bottom:28px">
              <thead><tr>
                <th {th}>Employee</th>
                <th {th} style="text-align:center">Total Hours Today</th>
              </tr></thead>
              <tbody>{e_total_rows}</tbody>
            </table>

            <h2 style="font-size:14px;font-weight:800;color:#0f172a;margin:0 0 10px">📋 Employee Breakdown</h2>
            <table width="100%" cellpadding="0" cellspacing="0"
              style="border-collapse:collapse;border:1px solid {bdr};border-radius:10px;overflow:hidden">
              <thead><tr>
                <th {th}>Employee</th>
                <th {th}>Project</th>
                <th {th}>Module</th>
                <th {th} style="text-align:center">Hours</th>
              </tr></thead>
              <tbody>{e_detail_rows}</tbody>
            </table>"""

        return f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f8fafc;font-family:Segoe UI,Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0"><tr><td style="padding:32px 16px">
<table width="600" align="center" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,.08);overflow:hidden;max-width:600px">
  <tr><td style="background:linear-gradient(135deg,{acc} 0%,{grn} 100%);padding:28px 36px">
    <div style="font-size:22px;font-weight:900;color:#fff;letter-spacing:-.5px">⏱ Pulse Tracker</div>
    <div style="color:rgba(255,255,255,.9);font-size:14px;margin-top:6px;font-weight:600">📊 Daily Summary — {date_str}</div>
  </td></tr>
  <tr><td style="padding:28px 36px 36px">{body}
    <p style="font-size:11px;color:#94a3b8;margin-top:32px;padding-top:16px;border-top:1px solid {bdr}">
      Automated daily summary · Sent every day at 08:30 AM · Pulse Tracker
    </p>
  </td></tr>
</table></td></tr></table></body></html>"""
    finally:
        try: store2.close()
        except: pass


def _send_daily_email(target_date=None):
    if not SMTP_USER or not SMTP_PASS:
        print("[Email] SMTP_USER/SMTP_PASS not set — skipping"); return
    if target_date is None:
        target_date = _date.today() - _td(days=1)
    html = _build_email_html(target_date)
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"Pulse Tracker — Daily Summary {target_date.strftime('%d %b %Y')}"
    msg["From"] = SMTP_FROM; msg["To"] = REPORT_TO
    msg.attach(MIMEText(html, "html", "utf-8"))
    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=20) as s:
            s.ehlo(); s.starttls(); s.ehlo()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(SMTP_FROM, [REPORT_TO], msg.as_string())
        print(f"[Email] Sent to {REPORT_TO} for {target_date}")
    except Exception as e:
        print(f"[Email] Error: {e}")


def _check_unpunched_alerts():
    """
    For every employee who has an open session (no punch_out) older than
    UNPUNCHED_HOURS, send an email to THAT employee (if email is set) and
    a CC to the manager (REPORT_TO).
    Runs every 30 minutes via the scheduler so alerts are timely.
    """
    if not os.environ.get("SMTP_USER") or not os.environ.get("SMTP_PASS"):
        return
    try:
        store2 = DataStore(DB_PATH)
        d2     = store2.load()
        now_dt = _dt.now()
        emp_map = {e.id: e for e in d2.employees}
        proj_map= {p.id: p for p in d2.projects}

        for s in d2.sessions:
            if s.punch_out:          continue   # already closed
            emp = emp_map.get(s.employee_id)
            if not emp:              continue

            # How long has this session been open?
            open_hours = (now_dt - s.punch_in).total_seconds() / 3600.0
            if open_hours < UNPUNCHED_HOURS:
                continue

            emp_email = getattr(emp, "email", "").strip()
            if not emp_email:
                continue  # no email on record — skip individual alert

            proj = proj_map.get(s.project_id)
            proj_name = proj.name if proj else s.project_id
            punch_in_str = s.punch_in.strftime("%d %b %Y at %I:%M %p")
            hours_str = "%dh %02dm" % (int(open_hours), int((open_hours % 1) * 60))

            subject = (f"[Pulse Tracker] Reminder: You have not punched out — "
                       f"{emp.name} · {punch_in_str}")

            html = f"""<!DOCTYPE html>
<html><body style="margin:0;padding:0;background:#f8fafc;font-family:Segoe UI,Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td style="padding:32px 16px">
<table width="560" align="center" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,.08);overflow:hidden;max-width:560px">
  <!-- Header -->
  <tr><td style="background:linear-gradient(135deg,#dc2626 0%,#b91c1c 100%);padding:28px 32px">
    <div style="font-size:22px;font-weight:800;color:#fff;letter-spacing:-0.5px">
      ⏱ Pulse Tracker
    </div>
    <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:6px">
      Session Still Open — Action Required
    </div>
  </td></tr>
  <!-- Body -->
  <tr><td style="padding:28px 32px">
    <p style="font-size:16px;font-weight:700;color:#1e293b;margin:0 0 8px">
      Hi {emp.name},
    </p>
    <p style="font-size:14px;color:#475569;margin:0 0 24px;line-height:1.6">
      Our records show you have an <strong>open session that has not been
      punched out</strong>. Please punch out as soon as possible or contact
      your manager to close it.
    </p>
    <!-- Session detail card -->
    <table width="100%" cellpadding="0" cellspacing="0"
      style="background:#fef2f2;border:1px solid #fecaca;border-radius:10px;margin-bottom:24px">
      <tr><td style="padding:18px 22px">
        <div style="font-size:11px;font-weight:700;color:#991b1b;text-transform:uppercase;
             letter-spacing:.8px;margin-bottom:12px">Open Session Details</div>
        <table width="100%" cellpadding="4">
          <tr>
            <td style="font-size:12px;color:#64748b;font-weight:600;width:130px">Project</td>
            <td style="font-size:13px;font-weight:700;color:#1e293b">{proj_name}</td>
          </tr>
          <tr>
            <td style="font-size:12px;color:#64748b;font-weight:600">Punched In</td>
            <td style="font-size:13px;font-weight:700;color:#1e293b">{punch_in_str}</td>
          </tr>
          <tr>
            <td style="font-size:12px;color:#64748b;font-weight:600">Duration Open</td>
            <td style="font-size:13px;font-weight:700;color:#dc2626">{hours_str}</td>
          </tr>
        </table>
      </td></tr>
    </table>
    <!-- CTA -->
    <div style="background:#fff7ed;border:1px solid #fed7aa;border-radius:10px;padding:16px 22px;margin-bottom:24px">
      <div style="font-size:13px;font-weight:700;color:#c2410c;margin-bottom:6px">
        ⚠️ What to do
      </div>
      <div style="font-size:13px;color:#78350f;line-height:1.7">
        1. Log in to <strong>Pulse Tracker</strong> and punch out immediately.<br>
        2. If you are unable to, please <strong>contact your manager</strong>
           at <a href="mailto:{REPORT_TO}" style="color:#c2410c">{REPORT_TO}</a>
           to manually close the session.
      </div>
    </div>
    <p style="font-size:11px;color:#94a3b8;padding-top:16px;border-top:1px solid #e2e8f0;margin:0">
      This is an automated alert from Pulse Tracker.
      Sent because your session has been open for more than {UNPUNCHED_HOURS} hours.
    </p>
  </td></tr>
</table>
</td></tr></table>
</body></html>"""

            # Send to employee + CC manager
            smtp_user = os.environ.get("SMTP_USER","")
            smtp_pass = os.environ.get("SMTP_PASS","")
            smtp_host = os.environ.get("SMTP_HOST","smtp.gmail.com")
            smtp_port = int(os.environ.get("SMTP_PORT","587"))
            smtp_from = os.environ.get("SMTP_FROM", smtp_user)

            import smtplib
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"]    = smtp_from
            msg["To"]      = emp_email
            msg["Cc"]      = REPORT_TO
            msg.attach(MIMEText(html, "html", "utf-8"))
            recipients = list({emp_email, REPORT_TO})
            try:
                with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as smtp:
                    smtp.ehlo(); smtp.starttls(); smtp.ehlo()
                    smtp.login(smtp_user, smtp_pass)
                    smtp.sendmail(smtp_from, recipients, msg.as_string())
                print(f"[Unpunched] Alert sent to {emp_email} (CC {REPORT_TO}) for {emp.name} — {hours_str} open")
            except Exception as mail_err:
                print(f"[Unpunched] Mail error for {emp.name}: {mail_err}")

    except Exception as e:
        print(f"[Unpunched] Check error: {e}")
        import traceback; traceback.print_exc()


def _build_weekly_digest_html(week_start: "_date", week_end: "_date") -> str:
    from datetime import datetime as _dt2
    store2 = DataStore(DB_PATH)
    try:
        d2   = store2.load()
        em   = {e.id: e for e in d2.employees}
        pm   = {p.id: p for p in d2.projects}

        proj_used_sec: dict = {}   # pid -> seconds used this week
        emp_h:         dict = {}   # emp name -> hours this week

        for s in d2.sessions:
            if not s.punch_out: continue
            sd = s.punch_in.date()
            if sd < week_start or sd > week_end: continue
            en  = (em.get(s.employee_id) or type("X",(),{"name":s.employee_id})()).name
            raw = max(0, int((s.punch_out - s.punch_in).total_seconds()))
            proj_used_sec[s.project_id] = proj_used_sec.get(s.project_id, 0) + raw
            emp_h[en] = emp_h.get(en, 0.0) + raw / 3600.0

        # Colours & helpers
        acc  = "#2563eb"; grn = "#16a34a"; red = "#dc2626"
        amb  = "#d97706"; bdr = "#e2e8f0"
        def hhmm(sec):
            sec = max(0, int(sec))
            return "%02d:%02d" % (sec // 3600, (sec % 3600) // 60)
        th = f'style="background:{acc};color:#fff;padding:10px 14px;text-align:left;font-size:12px;font-weight:700;border:none"'
        td = 'style="padding:10px 14px;border-bottom:1px solid #f1f5f9;font-size:13px;color:#374151;vertical-align:middle"'

        w_str   = f"{week_start.strftime('%d %b')} – {week_end.strftime('%d %b %Y')}"
        total_s = sum(proj_used_sec.values())

        # ── Project rows with full planned/used/remaining/% ───────────────────
        proj_rows = ""
        for p in sorted(d2.projects, key=lambda x: proj_used_sec.get(x.id, 0), reverse=True):
            used_s = proj_used_sec.get(p.id, 0)
            if used_s == 0: continue
            plan_s   = int((p.planned_hours or 0) * 3600)
            rem_s    = max(0, plan_s - used_s) if plan_s > 0 else None
            pct      = round(min(100, used_s / plan_s * 100)) if plan_s > 0 else None
            # progress bar
            if pct is not None:
                bar_col = red if pct >= 100 else (amb if pct >= 80 else grn)
                bar_w   = min(100, pct)
                bar = (f'<div style="background:#e2e8f0;border-radius:99px;height:7px;width:110px;display:inline-block;vertical-align:middle">' +
                       f'<div style="background:{bar_col};border-radius:99px;height:7px;width:{bar_w}%;"></div></div>' +
                       f' <span style="font-size:11px;font-weight:700;color:{bar_col}">{pct}%</span>')
            else:
                bar = '<span style="color:#cbd5e1;font-size:12px">No budget</span>'
            pct_col  = red if pct and pct >= 100 else (amb if pct and pct >= 80 else grn)
            rem_str  = hhmm(rem_s) if rem_s is not None else "—"
            rem_col  = red if (rem_s is not None and rem_s == 0 and plan_s > 0) else grn
            proj_rows += (
                f'<tr>' +
                f'<td {td}><strong style="color:#0f172a">{p.name}</strong>' +
                (f'<br><span style="font-size:11px;background:#e0e7ff;color:#3730a3;border-radius:4px;padding:1px 6px">{p.code}</span>' if p.code else '') +
                f'</td>' +
                f'<td {td} style="font-family:monospace;color:#374151;text-align:center">' +
                (hhmm(plan_s) if plan_s else '<span style="color:#cbd5e1">—</span>') +
                f'</td>' +
                f'<td {td} style="font-family:monospace;color:{acc};font-weight:700;text-align:center">{hhmm(used_s)}</td>' +
                f'<td {td} style="font-family:monospace;color:{rem_col};font-weight:700;text-align:center">{rem_str}</td>' +
                f'<td {td}>{bar}</td>' +
                f'</tr>'
            )

        # ── Employee rows ─────────────────────────────────────────────────────
        emp_rows = ""
        for en in sorted(emp_h, key=lambda x: emp_h[x], reverse=True):
            emp_rows += (f'<tr><td {td}><strong>{en}</strong></td>' +
                         f'<td {td} style="font-family:monospace;color:{grn};font-weight:700;text-align:center">{hhmm(int(emp_h[en]*3600))}</td></tr>')

        if not proj_rows:
            body = f'<p style="color:#94a3b8;text-align:center;padding:32px;font-size:14px">No sessions recorded this week.</p>'
        else:
            body = f"""
            <div style="display:flex;gap:16px;margin-bottom:28px;flex-wrap:wrap">
              <div style="flex:1;background:#eff6ff;border-radius:12px;padding:18px 14px;text-align:center;min-width:110px">
                <div style="font-size:26px;font-weight:900;color:{acc};font-family:monospace">{hhmm(total_s)}</div>
                <div style="font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:5px">Total Hours</div>
              </div>
              <div style="flex:1;background:#f0fdf4;border-radius:12px;padding:18px 14px;text-align:center;min-width:110px">
                <div style="font-size:26px;font-weight:900;color:{grn};font-family:monospace">{len(proj_used_sec)}</div>
                <div style="font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:5px">Active Projects</div>
              </div>
              <div style="flex:1;background:#fefce8;border-radius:12px;padding:18px 14px;text-align:center;min-width:110px">
                <div style="font-size:26px;font-weight:900;color:#d97706;font-family:monospace">{len(emp_h)}</div>
                <div style="font-size:10px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.5px;margin-top:5px">Contributors</div>
              </div>
            </div>
            <h2 style="font-size:14px;font-weight:800;color:#0f172a;margin:0 0 10px;letter-spacing:-.2px">📊 Project Summary</h2>
            <table width="100%" cellpadding="0" cellspacing="0"
              style="border-collapse:collapse;border:1px solid {bdr};border-radius:10px;overflow:hidden;margin-bottom:28px">
              <thead><tr>
                <th {th}>Project</th>
                <th {th} style="text-align:center">Planned</th>
                <th {th} style="text-align:center">Used</th>
                <th {th} style="text-align:center">Remaining</th>
                <th {th}>Budget Used</th>
              </tr></thead>
              <tbody>{proj_rows}</tbody>
            </table>
            <h2 style="font-size:14px;font-weight:800;color:#0f172a;margin:0 0 10px">👥 Team Contribution</h2>
            <table width="100%" cellpadding="0" cellspacing="0"
              style="border-collapse:collapse;border:1px solid {bdr};border-radius:10px;overflow:hidden">
              <thead><tr>
                <th {th}>Employee</th>
                <th {th} style="text-align:center">Total Hours This Week</th>
              </tr></thead>
              <tbody>{emp_rows}
                <tr style="background:#eff6ff">
                  <td style="padding:10px 14px;font-size:13px;font-weight:800;color:#0f172a;border-top:2px solid #bfdbfe">Grand Total</td>
                  <td style="padding:10px 14px;font-family:monospace;font-weight:900;color:#2563eb;font-size:14px;text-align:center;border-top:2px solid #bfdbfe">{hhmm(total_s)}</td>
                </tr>
              </tbody>
            </table>"""

        return f"""<!DOCTYPE html><html><body style="margin:0;padding:0;background:#f8fafc;font-family:Segoe UI,Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0"><tr><td style="padding:32px 16px">
<table width="600" align="center" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:16px;box-shadow:0 4px 24px rgba(0,0,0,.08);overflow:hidden;max-width:600px">
  <tr><td style="background:linear-gradient(135deg,{acc} 0%,{grn} 100%);padding:30px 36px">
    <div style="font-size:22px;font-weight:900;color:#fff;letter-spacing:-.5px">⏱ Pulse Tracker</div>
    <div style="color:rgba(255,255,255,.9);font-size:14px;margin-top:6px;font-weight:600">📅 Weekly Digest — {w_str}</div>
  </td></tr>
  <tr><td style="padding:28px 36px 36px">{body}
    <p style="font-size:11px;color:#94a3b8;margin-top:32px;padding-top:16px;border-top:1px solid {bdr}">
      Automated weekly digest · Sent every Monday at 08:30 AM · Pulse Tracker
    </p>
  </td></tr>
</table></td></tr></table></body></html>"""
    finally:
        try: store2.close()
        except: pass



def _check_weekly_digest():
    """Send weekly digest on Monday at 08:30."""
    from datetime import datetime as _dt2
    if not (os.environ.get("SMTP_USER") and os.environ.get("SMTP_PASS")):
        return
    today = _dt2.now().date()
    if today.weekday() != 0:   # 0 = Monday
        return
    week_end   = today - timedelta(days=1)          # Sunday
    week_start = week_end - timedelta(days=6)       # previous Monday
    html_body  = _build_weekly_digest_html(week_start, week_end)
    _send_email_raw(
        subject=f"📅 Pulse Tracker — Weekly Digest ({week_start.strftime('%d %b')}–{week_end.strftime('%d %b %Y')})",
        html=html_body
    )
    print(f"[Digest] Weekly digest sent for {week_start} → {week_end}")

def _auto_close_sessions():
    """
    Auto punch-out: close any sessions still open at AUTO_CLOSE_HOUR.
    Sets punch_out to AUTO_CLOSE_HOUR:00:00 of the same day the session started.
    If the session started on a PREVIOUS day, closes it at midnight of the start day.
    Runs every minute via the scheduler, fires once per day at AUTO_CLOSE_HOUR.
    """
    if not AUTO_CLOSE_ENABLED:
        return
    try:
        store2  = DataStore(DB_PATH)
        d2      = store2.load()
        now_dt  = _dt.now()
        closed  = 0

        for s in d2.sessions:
            if s.punch_out:
                continue  # already closed

            punch_in_dt = s.punch_in if isinstance(s.punch_in, _dt) else _dt.strptime(str(s.punch_in)[:19], "%Y-%m-%d %H:%M:%S")
            start_date  = punch_in_dt.date()
            today       = now_dt.date()

            # Case 1: session started today — close at AUTO_CLOSE_HOUR today
            if start_date == today:
                close_dt = punch_in_dt.replace(
                    hour=AUTO_CLOSE_HOUR, minute=0, second=0
                )
                # Only close if we've passed the auto-close hour
                if now_dt.hour < AUTO_CLOSE_HOUR:
                    continue

            # Case 2: session started on a PREVIOUS day — close at 23:59 of start day
            else:
                close_dt = punch_in_dt.replace(
                    hour=23, minute=59, second=0
                )

            # Sanity: close_dt must be after punch_in
            if close_dt <= punch_in_dt:
                close_dt = punch_in_dt + _td(seconds=1)

            # Calculate duration
            dur_td   = close_dt - punch_in_dt
            total_s  = int(dur_td.total_seconds())
            hours    = total_s // 3600
            minutes  = (total_s % 3600) // 60
            seconds  = total_s % 60
            duration = f"{hours}:{minutes:02d}:{seconds:02d}"

            # Update session
            s.punch_out = close_dt.strftime("%Y-%m-%d %H:%M:%S")
            s.duration  = duration
            closed += 1

        if closed > 0:
            store2.save(d2)
            print(f"[AutoClose] Closed {closed} open session(s) at {AUTO_CLOSE_HOUR}:00")

    except Exception as e:
        print(f"[AutoClose] Error: {e}")


def _scheduler_loop():
    sent_on          = None   # date: tracks 08:30 daily jobs
    unpunched_sent   = None   # date: tracks 23:55 unpunched alert
    auto_close_sent  = None   # date: tracks auto-close
    while True:
        _now = _dt.now()
        today = _now.date()

        # ── AUTO-CLOSE — close open sessions at configured hour ──────────────
        if AUTO_CLOSE_ENABLED and _now.hour >= AUTO_CLOSE_HOUR:
            if auto_close_sent != today:
                auto_close_sent = today
                try: _auto_close_sessions()
                except Exception as e: print(f"[Scheduler] auto_close: {e}")

        # ── 23:55 — unpunched-out alert (end of day) ──────────────────────────
        if (_now.hour == 23 and _now.minute >= 55):
            if unpunched_sent != today:
                unpunched_sent = today
                try: _check_unpunched_alerts()
                except Exception as e: print(f"[Scheduler] unpunched: {e}")

        # ── 08:30 — daily summary, weekly digest, inactivity ─────────────────
        if _now.hour > 8 or (_now.hour == 8 and _now.minute >= 30):
            if sent_on != today:
                sent_on = today
                try: _send_daily_email()
                except Exception as e: print(f"[Scheduler] daily: {e}")
                try: _check_weekly_digest()
                except Exception as e: print(f"[Scheduler] weekly: {e}")

        _time.sleep(60)

_sched_thread = threading.Thread(target=_scheduler_loop, daemon=True)
_sched_thread.start()
print("[Scheduler] Daily email at 08:30 → " + REPORT_TO)

@app.route("/api/admin/send-report", methods=["POST"])
@login_required
@admin_required
def manual_send_report():
    b = request.json or {}
    date_str = b.get("date") or ""
    from datetime import datetime as _dtt
    try:
        target = _dtt.strptime(date_str, "%Y-%m-%d").date() if date_str else (_date.today() - _td(days=1))
    except:
        target = _date.today() - _td(days=1)
    threading.Thread(target=_send_daily_email, args=(target,), daemon=True).start()
    return jsonify({"ok": True, "message": f"Report queued for {target} → {REPORT_TO}"})

@app.route("/api/admin/email-status", methods=["GET"])
@login_required
@admin_required
def email_status():
    return jsonify({
        "configured":     bool(SMTP_USER and SMTP_PASS),
        "smtp_host":      SMTP_HOST,
        "report_to":      REPORT_TO,
        "smtp_user":      SMTP_USER or "(not set)",
        "unpunched_hours": UNPUNCHED_HOURS,
    })



@app.route("/api/admin/email-config", methods=["GET"])
@login_required
@admin_required
def get_email_config():
    """Return full email config + live scheduler status for the settings page."""
    import datetime as _dt3
    from datetime import datetime as _dtx
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    configured = bool(smtp_user and smtp_pass)
    now = _dtx.now()
    # Next 08:30 trigger
    next_run = now.replace(hour=8, minute=30, second=0, microsecond=0)
    if now >= next_run:
        next_run = next_run + _td(days=1)
    next_monday = next_run
    while next_monday.weekday() != 0:
        next_monday += _td(days=1)
    return jsonify({
        "configured": configured,
        "smtp_host":  os.environ.get("SMTP_HOST", "smtp.gmail.com"),
        "smtp_port":  os.environ.get("SMTP_PORT", "587"),
        "smtp_user":  smtp_user or "",
        "smtp_from":  os.environ.get("SMTP_FROM", smtp_user),
        "report_to":  REPORT_TO,
        "inactivity_days": INACTIVITY_DAYS,
        "daily_time": "08:30 AM",
        "weekly_day": "Monday",
        "next_daily":   next_run.strftime("%d %b %Y %H:%M"),
        "next_weekly":  next_monday.strftime("%d %b %Y %H:%M"),
        "server_time":  now.strftime("%d %b %Y %H:%M:%S"),
        "checks": {
            "smtp_host_set":  bool(os.environ.get("SMTP_HOST")),
            "smtp_port_set":  bool(os.environ.get("SMTP_PORT")),
            "smtp_user_set":  bool(smtp_user),
            "smtp_pass_set":  bool(smtp_pass),
            "smtp_from_set":  bool(os.environ.get("SMTP_FROM")),
            "report_to_set":  bool(REPORT_TO),
        }
    })


@app.route("/api/admin/email-config", methods=["POST"])
@login_required
@admin_required
def save_email_config():
    """Persist SMTP settings as env vars for the current process (runtime only).
       For permanent config the user should set env vars before starting server."""
    b = request.json or {}
    fields = ["SMTP_HOST", "SMTP_PORT", "SMTP_USER", "SMTP_PASS", "SMTP_FROM", "REPORT_TO_ADDR"]
    for field in fields:
        val = (b.get(field) or "").strip()
        if val:
            os.environ[field] = val
    # Apply REPORT_TO immediately
    global REPORT_TO, SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS, SMTP_FROM, INACTIVITY_DAYS
    if b.get("REPORT_TO_ADDR"):
        REPORT_TO = b["REPORT_TO_ADDR"].strip()
    if b.get("SMTP_HOST"):  SMTP_HOST = b["SMTP_HOST"].strip()
    if b.get("SMTP_PORT"):  SMTP_PORT = int(b["SMTP_PORT"].strip())
    if b.get("SMTP_USER"):  SMTP_USER = b["SMTP_USER"].strip()
    if b.get("SMTP_PASS"):  SMTP_PASS = b["SMTP_PASS"].strip()
    if b.get("SMTP_FROM"):  SMTP_FROM = b["SMTP_FROM"].strip()
    if b.get("INACTIVITY_DAYS"):
        try: INACTIVITY_DAYS = int(b["INACTIVITY_DAYS"])
        except: pass
    return jsonify({"ok": True, "message": "Settings saved for this session."})


@app.route("/api/admin/email-test", methods=["POST"])
@login_required
@admin_required
def test_email_connection():
    """Test SMTP connection and send a test email."""
    import smtplib as _sm2
    try:
        h = os.environ.get("SMTP_HOST", SMTP_HOST) or "smtp.gmail.com"
        p = int(os.environ.get("SMTP_PORT", str(SMTP_PORT)) or 587)
        u = os.environ.get("SMTP_USER", SMTP_USER) or ""
        pw= os.environ.get("SMTP_PASS", SMTP_PASS) or ""
        fr= os.environ.get("SMTP_FROM", SMTP_FROM) or u
        to= REPORT_TO
        if not u or not pw:
            return jsonify({"ok": False, "error": "SMTP_USER and SMTP_PASS are required"}), 400
        # Step 1: connect
        with _sm2.SMTP(h, p, timeout=10) as s:
            s.ehlo()
            s.starttls()
            s.ehlo()
            # Step 2: login
            s.login(u, pw)
            # Step 3: send test email
            from email.mime.multipart import MIMEMultipart as _MM
            from email.mime.text import MIMEText as _MT
            msg = _MM("alternative")
            msg["Subject"] = "✅ Pulse Tracker — SMTP Test Successful"
            msg["From"] = fr; msg["To"] = to
            html = f"""<div style="font-family:Arial;padding:24px">
              <h2 style="color:#1E3A5F">✅ SMTP Connection Test Passed</h2>
              <p>Your Pulse Tracker email configuration is working correctly.</p>
              <table style="background:#f8fafc;border-radius:8px;padding:16px;width:100%">
                <tr><td style="color:#64748b;padding:4px 8px">SMTP Host</td><td><strong>{h}:{p}</strong></td></tr>
                <tr><td style="color:#64748b;padding:4px 8px">Auth User</td><td><strong>{u}</strong></td></tr>
                <tr><td style="color:#64748b;padding:4px 8px">Recipient</td><td><strong>{to}</strong></td></tr>
              </table>
              <p style="color:#64748b;font-size:12px;margin-top:16px">Sent from Pulse Tracker · {__import__("datetime").datetime.now().strftime("%d %b %Y %H:%M")}</p>
            </div>"""
            msg.attach(_MT(html, "html", "utf-8"))
            s.sendmail(fr, [to], msg.as_string())
        return jsonify({"ok": True, "message": f"Test email sent to {to}"})
    except _sm2.SMTPAuthenticationError:
        return jsonify({"ok": False, "error": "Authentication failed — check SMTP_USER and SMTP_PASS (Gmail: use App Password)"}), 400
    except _sm2.SMTPConnectError as e:
        return jsonify({"ok": False, "error": f"Cannot connect to {SMTP_HOST}:{SMTP_PORT} — {e}"}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ── Excel Session Import ───────────────────────────────────────────────────────
@app.route("/api/admin/import-sessions", methods=["POST"])
@login_required
@admin_required
def import_sessions():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "Only .xlsx files are supported"}), 400

    import openpyxl, io
    from datetime import datetime as _dt

    data = f.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        return jsonify({"error": f"Cannot read Excel file: {e}"}), 400

    # Find the "Session Import" sheet; fall back to first sheet
    ws = wb["Session Import"] if "Session Import" in wb.sheetnames else wb.active

    store  = get_store()
    d      = store.load()
    emp_by_code = {(e.emp_code or "").strip().upper(): e for e in d.employees}
    emp_by_name = {e.name.strip().lower(): e for e in d.employees}
    proj_by_code= {p.code.strip().upper(): p for p in d.projects}

    results  = []
    imported = 0
    skipped  = 0
    errors   = 0
    MAX_ROWS = 500

    # Detect header row (look for "Date" in col A between rows 1-10)
    header_row = None
    for rx in range(1, 11):
        v = ws.cell(row=rx, column=1).value
        if v and str(v).strip().lower() in ("date", "date *"):
            header_row = rx
            break
    if header_row is None:
        return jsonify({"error": "Cannot find header row. Make sure column A header says 'Date'."}), 400

    processed = 0
    for rx in range(header_row + 2, ws.max_row + 1):  # +2 skips sub-hint row
        if processed >= MAX_ROWS:
            break
        row_vals = [ws.cell(row=rx, column=cx).value for cx in range(1, 10)]
        # Skip completely blank rows
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        processed += 1

        def cell(idx):  # 0-based
            v = row_vals[idx]
            return str(v).strip() if v is not None else ""

        date_s    = cell(0)
        pi_time_s = cell(1)
        po_time_s = cell(2)
        emp_name  = cell(3)
        emp_code  = cell(4).upper()
        proj_code = cell(5).upper()
        module    = cell(6)
        task      = cell(7)
        remark    = cell(8)

        row_result = {"row": rx, "employee": emp_name or emp_code, "project": proj_code}

        # ── Validate required fields ───────────────────────────────────────
        missing = []
        if not date_s:    missing.append("Date")
        if not pi_time_s: missing.append("Punch In Time")
        if not emp_code and not emp_name: missing.append("Employee Code or Name")
        if not proj_code: missing.append("Project Code")
        if missing:
            row_result.update({"status": "error", "detail": f"Missing: {', '.join(missing)}"})
            results.append(row_result); errors += 1; continue

        # ── Parse date & times ─────────────────────────────────────────────
        try:
            # date_s might be "2025-03-04" or a datetime object from Excel
            if hasattr(date_s, 'date'):
                d_obj = date_s.date()
            else:
                from datetime import date as _date
                d_obj = _date.fromisoformat(str(date_s).strip()[:10])
        except Exception:
            row_result.update({"status": "error", "detail": f"Invalid date format: '{date_s}'. Use YYYY-MM-DD"})
            results.append(row_result); errors += 1; continue

        def parse_time(ts):
            ts = str(ts).strip()
            if ":" not in ts:
                return None
            parts = ts.split(":")
            try:
                hh, mm = int(parts[0]), int(parts[1])
                return _dt(d_obj.year, d_obj.month, d_obj.day, hh, mm)
            except Exception:
                return None

        punch_in  = parse_time(pi_time_s)
        punch_out = parse_time(po_time_s) if po_time_s else None

        if punch_in is None:
            row_result.update({"status": "error", "detail": f"Invalid punch-in time: '{pi_time_s}'. Use HH:MM"})
            results.append(row_result); errors += 1; continue

        if punch_out and punch_out <= punch_in:
            row_result.update({"status": "error", "detail": "Punch out must be after punch in"})
            results.append(row_result); errors += 1; continue

        # ── Resolve employee ───────────────────────────────────────────────
        emp = emp_by_code.get(emp_code) or emp_by_name.get(emp_name.lower())
        if not emp:
            row_result.update({"status": "error", "detail": f"Employee not found: code='{emp_code}' name='{emp_name}'"})
            results.append(row_result); errors += 1; continue

        # ── Resolve project ────────────────────────────────────────────────
        proj = proj_by_code.get(proj_code)
        if not proj:
            row_result.update({"status": "error", "detail": f"Project code not found: '{proj_code}'"})
            results.append(row_result); errors += 1; continue

        # ── Check duplicate (same employee + same punch_in) ────────────────
        pi_iso = punch_in.isoformat(timespec="seconds")
        cur = store.conn.cursor()
        cur.execute("SELECT id FROM sessions WHERE employee_id=? AND punch_in=?", (emp.id, pi_iso))
        if cur.fetchone():
            dur = ""
            if punch_out:
                sec = int((punch_out - punch_in).total_seconds())
                dur = f"  ({sec//3600:02d}:{(sec%3600)//60:02d})"
            row_result.update({"status": "skipped", "detail": f"Duplicate session{dur}"})
            results.append(row_result); skipped += 1; continue

        # ── Normalize module & task ────────────────────────────────────────
        from core import normalize_module, TASK_CATEGORIES, MODULE_CATEGORIES, short_id
        mod_clean  = normalize_module(module) if module else ""
        task_clean = task if task in TASK_CATEGORIES else (TASK_CATEGORIES[0] if task else "")

        # ── Insert session ─────────────────────────────────────────────────
        sess_id = "imp_" + short_id(10)
        cur.execute(
            "INSERT INTO sessions(id,employee_id,project_id,module,task_category,remark,punch_in,punch_out) VALUES (?,?,?,?,?,?,?,?)",
            (sess_id, emp.id, proj.id, mod_clean, task_clean, remark,
             pi_iso, punch_out.isoformat(timespec="seconds") if punch_out else None)
        )
        store.conn.commit()

        dur_str = ""
        if punch_out:
            sec = int((punch_out - punch_in).total_seconds())
            dur_str = f"{sec//3600:02d}:{(sec%3600)//60:02d}"

        row_result.update({
            "status": "imported",
            "detail": f"✓ {emp.name} → {proj.name}  {dur_str}",
            "duration": dur_str
        })
        results.append(row_result)
        imported += 1

    return jsonify({
        "ok": True,
        "imported": imported,
        "skipped": skipped,
        "errors": errors,
        "total_processed": processed,
        "results": results
    })


@app.route("/api/admin/export", methods=["POST"])
@login_required
@admin_required
def export_excel():
    b = request.json or {}
    df = parse_date(b.get("date_from") or "")
    dt = parse_date(b.get("date_to")   or "")
    if not df or not dt:
        return jsonify({"error": "date_from and date_to are required"}), 400
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    try:
        store = get_store()
        approved_ids = set(store.approved_employee_ids(df, dt))
        _export_all_xlsx(store, tmp.name, d_from=df, d_to=dt, approved_ids=approved_ids)
        return send_file(
            tmp.name,
            as_attachment=True,
            download_name=f"TimeReport_{fmt_date(df)}_to_{fmt_date(dt)}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/import-template", methods=["GET"])
@login_required
def download_import_template():
    path = os.path.join(app.static_folder, "pulse_import_template.xlsx")
    if not os.path.exists(path):
        return jsonify({"error": "Template not found"}), 404
    return send_file(path, as_attachment=True,
                     download_name="pulse_import_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 11 — PROJECT MILESTONES
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/admin/projects/<pid>/milestones", methods=["GET"])
@login_required
def get_milestones(pid):
    store = get_store()
    cur = store.conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS milestones(
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            target_date TEXT NOT NULL,
            soi_level INTEGER DEFAULT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            completed_at TEXT DEFAULT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
        )
    """)
    store.conn.commit()
    cur.execute("SELECT * FROM milestones WHERE project_id=? ORDER BY target_date ASC", (pid,))
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    return jsonify(rows)

@app.route("/api/admin/milestones", methods=["GET"])
@login_required
def get_all_milestones():
    store = get_store()
    cur = store.conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS milestones(
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            target_date TEXT NOT NULL,
            soi_level INTEGER DEFAULT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            completed_at TEXT DEFAULT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
        )
    """)
    store.conn.commit()
    cur.execute("""
        SELECT m.*, p.name as project_name, p.code as project_code
        FROM milestones m
        LEFT JOIN projects p ON m.project_id = p.id
        ORDER BY m.target_date ASC
    """)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    return jsonify(rows)

@app.route("/api/admin/projects/<pid>/milestones", methods=["POST"])
@login_required
@admin_required
def add_milestone(pid):
    data = request.json or {}
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"error": "Title required"}), 400
    target_date = (data.get("target_date") or "").strip()
    if not target_date:
        return jsonify({"error": "Target date required"}), 400
    store = get_store()
    cur = store.conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS milestones(
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            title TEXT NOT NULL,
            description TEXT DEFAULT '',
            target_date TEXT NOT NULL,
            soi_level INTEGER DEFAULT NULL,
            status TEXT NOT NULL DEFAULT 'pending',
            completed_at TEXT DEFAULT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
        )
    """)
    mid = "m_" + short_id(10)
    cur.execute(
        "INSERT INTO milestones(id,project_id,title,description,target_date,soi_level,status,created_at) VALUES (?,?,?,?,?,?,?,?)",
        (mid, pid, title, data.get("description",""), target_date,
         data.get("soi_level"), "pending", now().isoformat(timespec="seconds"))
    )
    store.conn.commit()
    return jsonify({"ok": True, "id": mid})

@app.route("/api/admin/milestones/<mid>", methods=["PUT"])
@login_required
@admin_required
def update_milestone(mid):
    data = request.json or {}
    store = get_store()
    cur = store.conn.cursor()
    status = data.get("status", "pending")
    completed_at = now().isoformat(timespec="seconds") if status == "completed" else None
    cur.execute("""
        UPDATE milestones SET title=?, description=?, target_date=?, soi_level=?,
        status=?, completed_at=? WHERE id=?
    """, (data.get("title",""), data.get("description",""), data.get("target_date",""),
          data.get("soi_level"), status, completed_at, mid))
    store.conn.commit()
    return jsonify({"ok": True})

@app.route("/api/admin/milestones/<mid>", methods=["DELETE"])
@login_required
@admin_required
def delete_milestone(mid):
    store = get_store()
    store.conn.execute("DELETE FROM milestones WHERE id=?", (mid,))
    store.conn.commit()
    return jsonify({"ok": True})

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 13 — PROJECT LOGBOOK
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/admin/projects/<pid>/logbook", methods=["GET"])
@login_required
def get_logbook(pid):
    store = get_store()
    cur = store.conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS logbook(
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            author_name TEXT NOT NULL,
            author_id TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
        )
    """)
    store.conn.commit()
    cur.execute("SELECT * FROM logbook WHERE project_id=? ORDER BY created_at DESC", (pid,))
    cols = [d[0] for d in cur.description]
    return jsonify([dict(zip(cols, r)) for r in cur.fetchall()])

@app.route("/api/admin/projects/<pid>/logbook", methods=["POST"])
@login_required
def add_logbook_entry(pid):
    data = request.json or {}
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"error": "Content required"}), 400
    store = get_store()
    cur = store.conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS logbook(
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL,
            author_name TEXT NOT NULL,
            author_id TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
        )
    """)
    lid = "l_" + short_id(10)
    author_name = session.get("name", "Unknown")
    author_id   = session.get("user", "")
    cur.execute(
        "INSERT INTO logbook(id,project_id,author_name,author_id,content,created_at) VALUES (?,?,?,?,?,?)",
        (lid, pid, author_name, author_id, content, now().isoformat(timespec="seconds"))
    )
    store.conn.commit()
    return jsonify({"ok": True, "id": lid})

@app.route("/api/admin/logbook/<lid>", methods=["DELETE"])
@login_required
@admin_required
def delete_logbook_entry(lid):
    store = get_store()
    store.conn.execute("DELETE FROM logbook WHERE id=?", (lid,))
    store.conn.commit()
    return jsonify({"ok": True})

# ═══════════════════════════════════════════════════════════════════════════════
# FEATURE 14 — WEEKLY DIGEST EMAIL (Mondays 08:30)
# FEATURE 16 — INACTIVITY ALERT (N days no sessions)
# Both bolt onto the existing scheduler thread via separate check functions
# ═══════════════════════════════════════════════════════════════════════════════

INACTIVITY_DAYS  = int(os.environ.get("INACTIVITY_DAYS", "3"))
UNPUNCHED_HOURS  = int(os.environ.get("UNPUNCHED_HOURS", "10"))  # hours before unpunched-out alert

def _scheduler_loop():
    sent_on          = None   # date: tracks 08:30 daily jobs
    unpunched_sent   = None   # date: tracks 23:55 unpunched alert
    while True:
        _now = _dt.now()
        today = _now.date()

        # ── 23:55 — unpunched-out alert (end of day) ──────────────────────────
        if (_now.hour == 23 and _now.minute >= 55):
            if unpunched_sent != today:
                unpunched_sent = today
                try: _check_unpunched_alerts()
                except Exception as e: print(f"[Scheduler] unpunched: {e}")

        # ── 08:30 — daily summary, weekly digest, inactivity ─────────────────
        if _now.hour > 8 or (_now.hour == 8 and _now.minute >= 30):
            if sent_on != today:
                sent_on = today
                try: _send_daily_email()
                except Exception as e: print(f"[Scheduler] daily: {e}")
                try: _check_weekly_digest()
                except Exception as e: print(f"[Scheduler] weekly: {e}")

        _time.sleep(60)


# ── Serve SPA ─────────────────────────────────────────────────────────────────
from flask import send_file as _send_file
@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve(path):
    import os as _os
    sd = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "static")
    if path:
        f = _os.path.join(sd, path)
        if _os.path.isfile(f):
            return _send_file(f)
    return _send_file(_os.path.join(sd, "index.html"))

@app.route("/api/admin/task-categories", methods=["POST"])
@login_required
@admin_required
def add_task_category():
    """Add a new global task category and persist it in SQLite."""
    b    = request.json or {}
    name = (b.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    try:
        task_categories = get_store().add_task_category(name)
        return jsonify({"ok": True, "task_categories": task_categories})
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

@app.route("/api/admin/task-categories/<path:name>", methods=["DELETE"])
@login_required
@admin_required
def delete_task_category(name):
    """Delete a global task category and clean project mappings."""
    clean = (name or "").strip()
    if not clean:
        return jsonify({"error": "Task name required"}), 400
    try:
        task_categories = get_store().delete_task_category(clean)
        return jsonify({"ok": True, "task_categories": task_categories})
    except ValueError as e:
        return jsonify({"error": str(e)}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 400


if __name__ == "__main__":
    port = int(__import__('os').environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port, threaded=True)